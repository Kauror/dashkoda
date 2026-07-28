"""Builds synthetic canonical workbooks for the tests.

Every value here is invented. Nothing resembles Chamber legal work, and the
real workbook is never copied into the repository or into a fixture.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from apps.legal_work.workbook import (
    CONTROL_SHEET,
    DATA_COLUMNS,
    DATA_SHEET,
    DATA_TABLE_NAME,
    OVERVIEW_SHEET,
    WARNINGS_SHEET,
)

WARNINGS_TABLE_NAME = "tbl_oigusloome_warnings"
WARNINGS_COLUMNS = (
    "record_id",
    "source_row",
    "field",
    "warning_code",
    "original_value",
    "explanation",
)

# Dates are relative to today and always in the past. A fixed future year would
# be filtered out by the "received date must not be in the future" rule, which
# is exactly the behaviour the selectors are supposed to have.
TODAY = dt.date.today()
REPORTING_DATE = TODAY - dt.timedelta(days=1)
GENERATED_AT = dt.datetime.combine(REPORTING_DATE, dt.time(6, 30))

DEFAULT_RECEIVED = TODAY - dt.timedelta(days=60)
DEFAULT_DEADLINE = TODAY - dt.timedelta(days=30)


def synthetic_row(
    *,
    record_id: str,
    source_year: int = 2099,
    source_nr: int = 1,
    topic: str = "Sünteetiline testteema",
    act_type: str = "Sünteetiline liik",
    received_date: dt.date | None = DEFAULT_RECEIVED,
    deadline_date: dt.date | None = DEFAULT_DEADLINE,
    sent_date: dt.date | None = None,
    sent_status: str = "pending",
    recipient: str = "Sünteetiline saaja",
    stage: str = "sünteetiline seis",
    stage_key: str = "sünteetiline seis",
    next_step: str = "",
    is_open: bool = True,
    warning_codes: str | None = None,
    source_row: int = 2,
    refreshed_at: dt.datetime | None = GENERATED_AT,
) -> list:
    """One DATA row in canonical column order."""
    return [
        record_id,
        source_year,
        source_nr,
        topic,
        act_type,
        received_date,
        deadline_date,
        sent_date,
        sent_status,
        recipient,
        stage,
        stage_key,
        next_step,
        is_open,
        warning_codes,
        source_row,
        refreshed_at,
    ]


def default_rows() -> list[list]:
    """A small, deliberately varied set covering the selector rules."""
    return [
        synthetic_row(
            record_id="SYN-0001",
            topic="Sünteetiline avatud teema",
            received_date=TODAY - dt.timedelta(days=60),
            is_open=True,
            source_row=2,
        ),
        synthetic_row(
            record_id="SYN-0002",
            topic="Sünteetiline saadetud teema",
            received_date=TODAY - dt.timedelta(days=50),
            sent_date=TODAY - dt.timedelta(days=20),
            sent_status="sent",
            stage="jõustunud",
            stage_key="jõustunud",
            is_open=False,
            source_row=3,
        ),
        synthetic_row(
            record_id="SYN-0003",
            topic="Sünteetiline saatmata teema",
            received_date=TODAY - dt.timedelta(days=45),
            sent_status="not_sent",
            is_open=True,
            warning_codes="missing_stage",
            stage="",
            stage_key="",
            source_row=4,
        ),
    ]


def control_values(
    rows: list[list],
    *,
    dataset_key: str = "oigusloome",
    schema_version: str = "1.1",
    overrides: dict | None = None,
) -> list[tuple]:
    """CONTROL rows, including the banner lines the real generator writes."""
    is_open_index = DATA_COLUMNS.index("is_open")
    sent_status_index = DATA_COLUMNS.index("sent_status")
    warnings_index = DATA_COLUMNS.index("warning_codes")

    values = {
        "dataset_key": dataset_key,
        "schema_version": schema_version,
        "source_file_name": "sünteetiline-lähtefail.xlsx",
        "source_sheet": "2099",
        "source_modified_at": "2099-03-01T06:00:00+02:00",
        "source_sha256": "0" * 64,
        "generated_at": GENERATED_AT.isoformat(),
        "reporting_date": REPORTING_DATE,
        "total_record_count": len(rows),
        "open_record_count": sum(1 for row in rows if row[is_open_index] is True),
        "sent_record_count": sum(1 for row in rows if row[sent_status_index] == "sent"),
        "not_sent_record_count": sum(1 for row in rows if row[sent_status_index] == "not_sent"),
        "warning_record_count": sum(1 for row in rows if row[warnings_index]),
        "refresh_status": "completed_with_warnings",
        "generator_version": "1.1.1",
    }
    values.update(overrides or {})

    return [
        ("DASHKODA ÕIGUSLOOME ANDMEVOOG", None),
        *values.items(),
        ("This workbook is generated. Do not edit DATA manually.", None),
    ]


def write_workbook(
    path: Path,
    *,
    rows: list[list] | None = None,
    control_overrides: dict | None = None,
    dataset_key: str = "oigusloome",
    schema_version: str = "1.1",
    sheets: tuple[str, ...] | None = None,
    data_table_name: str = DATA_TABLE_NAME,
    columns: tuple[str, ...] = DATA_COLUMNS,
    formula_in_data: bool = False,
) -> Path:
    """Write a synthetic workbook, optionally broken in one specific way."""
    rows = default_rows() if rows is None else rows
    sheets = sheets or (CONTROL_SHEET, OVERVIEW_SHEET, DATA_SHEET, WARNINGS_SHEET)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for name in sheets:
        workbook.create_sheet(name)

    if CONTROL_SHEET in sheets:
        control = workbook[CONTROL_SHEET]
        for key, value in control_values(
            rows,
            dataset_key=dataset_key,
            schema_version=schema_version,
            overrides=control_overrides,
        ):
            control.append([key, value])

    if OVERVIEW_SHEET in sheets:
        # Formatted for people and never read by the importer.
        workbook[OVERVIEW_SHEET].append(["Sünteetiline ülevaade"])

    if DATA_SHEET in sheets:
        data = workbook[DATA_SHEET]
        data.append(list(columns))
        for row in rows:
            data.append(row)
        if formula_in_data and rows:
            data.cell(row=2, column=1).value = '=CONCATENATE("SYN","-0001")'
        if data_table_name:
            last_column = chr(ord("A") + len(columns) - 1)
            data.add_table(
                Table(
                    displayName=data_table_name,
                    ref=f"A1:{last_column}{len(rows) + 1}",
                )
            )

    if WARNINGS_SHEET in sheets:
        warnings_sheet = workbook[WARNINGS_SHEET]
        warnings_sheet.append(list(WARNINGS_COLUMNS))
        warning_index = DATA_COLUMNS.index("warning_codes")
        record_index = DATA_COLUMNS.index("record_id")
        source_row_index = DATA_COLUMNS.index("source_row")
        for row in rows:
            if not row[warning_index]:
                continue
            for code in str(row[warning_index]).split(";"):
                warnings_sheet.append(
                    [
                        row[record_index],
                        row[source_row_index],
                        "stage",
                        code.strip(),
                        "sünteetiline algväärtus",
                        "Sünteetiline selgitus.",
                    ]
                )
        warnings_sheet.add_table(
            Table(
                displayName=WARNINGS_TABLE_NAME,
                ref=f"A1:F{warnings_sheet.max_row}",
            )
        )

    workbook.save(path)
    return path
