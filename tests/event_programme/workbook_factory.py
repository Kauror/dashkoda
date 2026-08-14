"""Builds synthetic event-programme workbooks for the tests.

Every value here is invented. Nothing resembles a Chamber event, and the real
export is never copied into the repository or into a fixture.

The layout deliberately reproduces two awkward properties of the real
generator, because they are exactly what the parser has to cope with: each
DASH_* table sits below a two-row banner, and the CONTROL sheet is a text
key/value table whose counts arrive as strings.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from apps.event_programme.workbook import (
    CONTROL_SHEET,
    CONTROL_TABLE_NAME,
    EVENTS_COLUMNS,
    EVENTS_SHEET,
    EVENTS_TABLE_NAME,
    OCCURRENCES_SHEET,
    REVIEW_SHEET,
    TAG_MAP_SHEET,
    URL_OVERRIDES_SHEET,
)

# The banner rows the generator writes for the people who maintain the
# operational workbook. The tables start immediately below them.
BANNER_ROWS = 2
CONTROL_BANNER_ROWS = 4

EXPORT_REFRESHED_AT = "2099-01-02T06:30:00+02:00"

DEFAULT_START = dt.datetime(2099, 3, 4)

# Distinguishes "the caller did not mention an end date" from "the caller asked
# for no end date". The first means a single-day event, the second an undated one.
SAME_DAY = object()

# Estonian month names, as the Chamber's generator writes `event_month_label`.
MONTH_LABELS = (
    "jaanuar",
    "veebruar",
    "märts",
    "aprill",
    "mai",
    "juuni",
    "juuli",
    "august",
    "september",
    "oktoober",
    "november",
    "detsember",
)

# Passenger sheets: required to be present, never read as data. One column each
# is enough to carry a valid Excel Table.
PASSENGER_SHEETS = {
    OCCURRENCES_SHEET: ("tbl_dash_event_occurrences", ("event_id", "occurrence_number")),
    REVIEW_SHEET: ("tbl_dash_review", ("severity", "issue_code")),
    TAG_MAP_SHEET: ("tbl_dash_tag_map", ("short_name_normalized", "tag_key")),
    URL_OVERRIDES_SHEET: ("tbl_dash_url_overrides", ("service_code", "public_url")),
}


def synthetic_row(
    *,
    event_id: str,
    service_code: str,
    event_name: str = "Sünteetiline sündmus",
    start_date: dt.datetime | None = DEFAULT_START,
    end_date=SAME_DAY,
    event_status: str = "past",
    tag_key: str = "seminar",
    tag_label: str = "Seminar",
    event_type_key: str = "seminar",
    event_type_label: str = "Seminar",
    delivery_mode: str = "onsite",
    include_status: str = "YES",
    public_url: str | None = None,
    public_link_status: str = "not_linked",
    source_year: int = 2099,
    source_sheet: str = "KOOD 2099",
    source_row: int = 2,
    source_occurrence_count: int = 1,
    date_parse_status: str = "parsed_single",
    review_required: bool = False,
    warning_codes: str | None = None,
    price_status: str = "paid",
    member_price_eur: float | None = 100,
    nonmember_price_eur: float | None = 200,
    added_date: dt.datetime | None = None,
    planning_lead_days: int | None = None,
) -> dict:
    """One synthetic DASH_EVENTS row, keyed by column name.

    The calendar fields derive from `start_date` exactly as the generator
    derives them, so an undated row leaves all four empty rather than carrying
    a year with no date, and a March event really does carry March and Q1.

    `planning_lead_days` derives from the two dates the same way the real
    generator derives it, unless a caller states one explicitly — which is how a
    test builds the disagreeing row that proves the importer stores the source's
    figure rather than recomputing its own.
    """
    dated = start_date is not None
    if end_date is SAME_DAY:
        end_date = start_date
    if planning_lead_days is None and dated and added_date is not None:
        planning_lead_days = (start_date.date() - added_date.date()).days
    return {
        "event_id": event_id,
        "service_code": service_code,
        "event_name": event_name,
        "event_name_raw": event_name,
        "date_text_raw": "04.03.2099" if dated else "kevad",
        "start_date": start_date,
        "end_date": end_date if dated else None,
        "event_year": start_date.year if dated else None,
        "event_month": start_date.month if dated else None,
        "event_month_key": f"{start_date:%Y-%m}" if dated else None,
        "event_month_label": MONTH_LABELS[start_date.month - 1] if dated else None,
        "event_quarter": f"Q{(start_date.month - 1) // 3 + 1}" if dated else None,
        "event_status": event_status,
        "short_name_raw": "SÜN",
        "short_name_normalized": "syn",
        "tag_key": tag_key,
        "tag_label": tag_label,
        "event_type_key": event_type_key,
        "event_type_label": event_type_label,
        "delivery_mode": delivery_mode,
        "include_status": include_status,
        # Internal group columns exist in the file and must never reach a model
        # field: their business meaning has never been established.
        "group_raw": None,
        "group_secondary_raw": None,
        # The normalised price pair and the status are stored. The `*_raw`
        # echoes, the later-price pair and the discount columns are not.
        "member_price_raw": None if member_price_eur is None else str(member_price_eur),
        "member_price_eur": member_price_eur,
        "nonmember_price_raw": None if nonmember_price_eur is None else str(nonmember_price_eur),
        "nonmember_price_eur": nonmember_price_eur,
        "later_member_price_raw": None,
        "later_member_price_eur": None,
        "later_nonmember_price_raw": None,
        "later_nonmember_price_eur": None,
        "price_status": price_status,
        "discount_code": None,
        "discount_raw": None,
        "added_date": added_date,
        "planning_lead_days": planning_lead_days,
        "public_url": public_url,
        "public_link_status": public_link_status,
        "source_year": source_year,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "source_occurrence_count": source_occurrence_count,
        "date_parse_status": date_parse_status,
        "review_required": review_required,
        "warning_codes": warning_codes,
        "export_refreshed_at": EXPORT_REFRESHED_AT,
    }


def default_rows() -> list[dict]:
    """Three events: dated and linked, dated and unlinked, undated."""
    return [
        synthetic_row(
            event_id="EVENT-9001",
            service_code="9001",
            public_url="https://www.koda.ee/et/sundmused/synteetiline",
            public_link_status="linked_embedded_latest",
        ),
        synthetic_row(event_id="EVENT-9002", service_code="9002", source_row=3),
        synthetic_row(
            event_id="EVENT-9003",
            service_code="9003",
            source_row=4,
            start_date=None,
            end_date=None,
            event_status="date_unknown",
            date_parse_status="unparsed",
            review_required=True,
            warning_codes="date_unparsed",
        ),
    ]


def default_control(rows: list[dict]) -> dict[str, str]:
    """CONTROL that agrees with `rows`.

    Every value is a string, because that is what the Office Script writes.
    """
    linked = sum(1 for row in rows if row["public_url"])
    return {
        "dataset_key": "events",
        "schema_version": "1.0",
        "generator_name": "Sünteetiline generaator",
        "generator_version": "1.0.0",
        "source_workbook_name": "Sünteetiline lähtefail.xlsx",
        "refresh_status": "ready_with_warnings",
        "canonical_event_count": str(len(rows)),
        "qualifying_occurrence_count": str(len(rows)),
        "excluded_event_count": "0",
        "repeated_service_code_count": "0",
        "linked_public_url_count": str(linked),
        "distinct_short_name_count": "1",
        "blocking_error_count": "0",
        "warning_count": "1",
        "export_refreshed_at": EXPORT_REFRESHED_AT,
        "last_successful_refresh_at": EXPORT_REFRESHED_AT,
    }


def _write_table(sheet, *, table_name: str, columns, rows, banner_rows: int) -> None:
    for index in range(banner_rows):
        sheet.cell(row=index + 1, column=1, value=f"Sünteetiline bänner {index + 1}")

    header_row = banner_rows + 1
    for position, name in enumerate(columns, start=1):
        sheet.cell(row=header_row, column=position, value=name)

    for offset, values in enumerate(rows, start=header_row + 1):
        for position, value in enumerate(values, start=1):
            sheet.cell(row=offset, column=position, value=value)

    last_row = header_row + max(len(rows), 1)
    last_column = sheet.cell(row=header_row, column=len(columns)).column_letter
    sheet.add_table(Table(displayName=table_name, ref=f"A{header_row}:{last_column}{last_row}"))


def build_workbook(
    path: Path | str,
    *,
    rows: list[dict] | None = None,
    control: dict[str, str] | None = None,
    events_columns=EVENTS_COLUMNS,
    omit_sheets: tuple[str, ...] = (),
) -> Path:
    """Write a synthetic workbook and return its path."""
    rows = default_rows() if rows is None else rows
    control = default_control(rows) if control is None else control

    workbook = Workbook()
    workbook.remove(workbook.active)

    if CONTROL_SHEET not in omit_sheets:
        sheet = workbook.create_sheet(CONTROL_SHEET)
        _write_table(
            sheet,
            table_name=CONTROL_TABLE_NAME,
            columns=("key", "value"),
            rows=[(key, value) for key, value in control.items()],
            banner_rows=CONTROL_BANNER_ROWS,
        )

    if EVENTS_SHEET not in omit_sheets:
        sheet = workbook.create_sheet(EVENTS_SHEET)
        _write_table(
            sheet,
            table_name=EVENTS_TABLE_NAME,
            columns=events_columns,
            rows=[tuple(row.get(name) for name in events_columns) for row in rows],
            banner_rows=BANNER_ROWS,
        )

    for name, (table_name, columns) in PASSENGER_SHEETS.items():
        if name in omit_sheets:
            continue
        sheet = workbook.create_sheet(name)
        _write_table(
            sheet,
            table_name=table_name,
            columns=columns,
            rows=[],
            banner_rows=BANNER_ROWS,
        )

    destination = Path(path)
    workbook.save(destination)
    workbook.close()
    return destination
