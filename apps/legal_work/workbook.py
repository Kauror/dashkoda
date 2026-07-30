"""Deterministic reader for the canonical legal-work workbook.

This module knows the workbook contract and nothing about Django models, the
import registry or Microsoft Graph. It either returns a fully validated
:class:`ParsedWorkbook` or raises :class:`WorkbookContractError`.

The authoritative input is the Excel Table ``tbl_oigusloome`` on the ``DATA``
sheet. The ``OVERVIEW`` sheet is a human convenience inside the workbook and is
never read as a data source: it is formatted for people, and deriving records
from cell positions would break the moment someone adjusted the layout.

Nothing here interprets the legal meaning of a record. Values are parsed,
type-checked and rejected — never repaired, inferred or summarised.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

DATASET_KEY = "oigusloome"

# The workbook's own declared contract version. 1.0 and 1.1 describe an
# identical DATA table; 1.1 only added CONTROL metadata and multi-year support,
# so both are accepted and the value is recorded on the snapshot.
SUPPORTED_SCHEMA_VERSIONS = ("1.0", "1.1")

CONTROL_SHEET = "CONTROL"
OVERVIEW_SHEET = "OVERVIEW"
DATA_SHEET = "DATA"
WARNINGS_SHEET = "WARNINGS"
REQUIRED_SHEETS = (CONTROL_SHEET, OVERVIEW_SHEET, DATA_SHEET, WARNINGS_SHEET)

DATA_TABLE_NAME = "tbl_oigusloome"

# Exact names, in exact order. A reordered or renamed column is a contract
# break, not something to be matched up by guesswork.
DATA_COLUMNS = (
    "record_id",
    "source_year",
    "source_nr",
    "topic",
    "act_type",
    "received_date",
    "deadline_date",
    "sent_date",
    "sent_status",
    "recipient",
    "stage",
    "stage_key",
    "next_step",
    "is_open",
    "warning_codes",
    "source_row",
    "refreshed_at",
)

REQUIRED_CONTROL_KEYS = (
    "dataset_key",
    "schema_version",
    "source_file_name",
    "source_sheet",
    "source_modified_at",
    "source_sha256",
    "generated_at",
    "reporting_date",
    "total_record_count",
    "open_record_count",
    "sent_record_count",
    "not_sent_record_count",
    "warning_record_count",
    "refresh_status",
    "generator_version",
)

VALID_SENT_STATUSES = frozenset({"pending", "sent", "not_sent", "invalid"})

# CONTROL keys that must be *present* but whose value may legitimately be empty.
# `source_modified_at` is the modification time of the lawyers' operational file.
# A generator that reads the operational workbook from the cloud rather than from
# a locally synchronised copy has no such time to report, which is why both the
# parsed control record and the snapshot field are nullable. Every other required
# key must carry a real value.
OPTIONAL_VALUE_CONTROL_KEYS = frozenset({"source_modified_at"})

# The generator separates codes with ";"; "," is tolerated defensively because
# both have appeared in hand-checked exports.
WARNING_CODE_SEPARATORS = (";", ",")

MAX_TOPIC_LENGTH = 4000
MAX_SHORT_TEXT_LENGTH = 300


class WorkbookContractError(ValueError):
    """The workbook does not satisfy the agreed contract."""


@dataclass(frozen=True)
class WorkbookControl:
    dataset_key: str
    schema_version: str
    source_sheet: str
    source_modified_at: dt.datetime | None
    generated_at: dt.datetime
    reporting_date: dt.date
    total_record_count: int
    open_record_count: int
    sent_record_count: int
    not_sent_record_count: int
    warning_record_count: int
    refresh_status: str
    generator_version: str


@dataclass(frozen=True)
class WorkbookRow:
    record_id: str
    source_year: int
    source_nr: int | None
    topic: str
    act_type: str
    received_date: dt.date | None
    deadline_date: dt.date | None
    sent_date: dt.date | None
    sent_status: str
    recipient: str
    stage: str
    stage_key: str
    next_step: str
    is_open: bool
    warning_codes: list[str]
    source_row: int
    refreshed_at: dt.datetime | None


@dataclass(frozen=True)
class ParsedWorkbook:
    control: WorkbookControl
    rows: tuple[WorkbookRow, ...]
    warnings: tuple[dict, ...] = field(default=())

    @property
    def open_count(self) -> int:
        return sum(1 for row in self.rows if row.is_open)

    @property
    def sent_count(self) -> int:
        return sum(1 for row in self.rows if row.sent_status == "sent")

    @property
    def not_sent_count(self) -> int:
        return sum(1 for row in self.rows if row.sent_status == "not_sent")

    @property
    def warning_record_count(self) -> int:
        """Records carrying at least one code.

        Deliberately not the number of rows on the WARNINGS sheet: that sheet
        holds one row per individual warning, so a record with two warnings
        appears twice there but counts once here.
        """
        return sum(1 for row in self.rows if row.warning_codes)


def _text(value, *, limit: int = MAX_SHORT_TEXT_LENGTH) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:limit]


def _as_date(value, *, column: str, row_number: int) -> dt.date | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise WorkbookContractError(
        f"Reas {row_number} ei ole veerg {column!r} kuupäev vaid {type(value).__name__}."
    )


def _as_int(value, *, column: str, row_number: int, allow_none: bool = False) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if allow_none:
            return None
        raise WorkbookContractError(f"Reas {row_number} puudub kohustuslik veerg {column!r}.")
    if isinstance(value, bool) or not isinstance(value, int):
        raise WorkbookContractError(
            f"Reas {row_number} ei ole veerg {column!r} täisarv vaid {type(value).__name__}."
        )
    return value


def _as_bool(value, *, column: str, row_number: int) -> bool:
    if isinstance(value, bool):
        return value
    raise WorkbookContractError(
        f"Reas {row_number} ei ole veerg {column!r} tõeväärtus vaid {type(value).__name__}."
    )


def _split_warning_codes(value) -> list[str]:
    if value is None:
        return []
    text = str(value)
    for separator in WARNING_CODE_SEPARATORS[1:]:
        text = text.replace(separator, WARNING_CODE_SEPARATORS[0])
    return [part.strip() for part in text.split(WARNING_CODE_SEPARATORS[0]) if part.strip()]


def _parse_timestamp(value, *, key: str) -> dt.datetime | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, str):
        try:
            return dt.datetime.fromisoformat(value.strip())
        except ValueError as error:
            raise WorkbookContractError(f"CONTROL väli {key!r} ei ole kehtiv ajatempel.") from error
    raise WorkbookContractError(f"CONTROL väli {key!r} ei ole ajatempel.")


def _read_control(sheet) -> dict:
    """Collect the key/value pairs from CONTROL.

    A row with no key at all is the workbook's own banner or footnote line and is
    skipped. A row that *has* a key keeps it even when the value cell is empty:
    "the generator did not write this field" and "the generator wrote this field
    and it is legitimately empty" are different facts, and only
    :func:`_require_control` decides which keys may be empty. Unknown keys are
    ignored rather than rejected, so the generator can add metadata without
    breaking this importer.
    """
    values: dict = {}
    for key_cell, value_cell in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if key_cell is None:
            continue
        key = str(key_cell).strip()
        if key:
            values[key] = value_cell
    return values


def _is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _require_control(values: dict) -> WorkbookControl:
    missing = [key for key in REQUIRED_CONTROL_KEYS if key not in values]
    if missing:
        raise WorkbookContractError(f"CONTROL lehel puuduvad väljad: {', '.join(sorted(missing))}.")

    blank = [
        key
        for key in REQUIRED_CONTROL_KEYS
        if key not in OPTIONAL_VALUE_CONTROL_KEYS and _is_blank(values[key])
    ]
    if blank:
        raise WorkbookContractError(f"CONTROL lehe väljad on tühjad: {', '.join(sorted(blank))}.")

    dataset_key = _text(values["dataset_key"])
    if dataset_key != DATASET_KEY:
        raise WorkbookContractError(
            f"Vale andmestik: oodati {DATASET_KEY!r}, sain {dataset_key!r}."
        )

    schema_version = _text(values["schema_version"])
    if schema_version not in SUPPORTED_SCHEMA_VERSIONS:
        raise WorkbookContractError(
            f"Toetamata skeemi versioon {schema_version!r}. "
            f"Toetatud: {', '.join(SUPPORTED_SCHEMA_VERSIONS)}."
        )

    generated_at = _parse_timestamp(values["generated_at"], key="generated_at")
    if generated_at is None:
        raise WorkbookContractError("CONTROL väli 'generated_at' on kohustuslik.")

    reporting_raw = values["reporting_date"]
    if isinstance(reporting_raw, dt.datetime):
        reporting_date = reporting_raw.date()
    elif isinstance(reporting_raw, dt.date):
        reporting_date = reporting_raw
    else:
        parsed = _parse_timestamp(reporting_raw, key="reporting_date")
        if parsed is None:
            raise WorkbookContractError("CONTROL väli 'reporting_date' on kohustuslik.")
        reporting_date = parsed.date()

    counts = {}
    for key in (
        "total_record_count",
        "open_record_count",
        "sent_record_count",
        "not_sent_record_count",
        "warning_record_count",
    ):
        raw = values[key]
        if isinstance(raw, bool) or not isinstance(raw, int):
            raise WorkbookContractError(f"CONTROL väli {key!r} peab olema täisarv.")
        if raw < 0:
            raise WorkbookContractError(f"CONTROL väli {key!r} ei tohi olla negatiivne.")
        counts[key] = raw

    return WorkbookControl(
        dataset_key=dataset_key,
        schema_version=schema_version,
        source_sheet=_text(values["source_sheet"]),
        source_modified_at=_parse_timestamp(values["source_modified_at"], key="source_modified_at"),
        generated_at=generated_at,
        reporting_date=reporting_date,
        refresh_status=_text(values["refresh_status"]),
        generator_version=_text(values["generator_version"]),
        **counts,
    )


def _require_data_table(sheet) -> None:
    tables = dict(getattr(sheet, "tables", {}) or {})
    if DATA_TABLE_NAME not in tables:
        found = ", ".join(sorted(tables)) or "(ühtegi)"
        raise WorkbookContractError(
            f"DATA lehel puudub Exceli tabel {DATA_TABLE_NAME!r}. Leitud: {found}."
        )


def _require_header(sheet) -> None:
    header = tuple(_text(cell.value, limit=200) for cell in sheet[1][: len(DATA_COLUMNS)])
    if header != DATA_COLUMNS:
        raise WorkbookContractError(
            "DATA veerud ei vasta kokkuleppele.\n"
            f"Oodatud: {', '.join(DATA_COLUMNS)}\n"
            f"Leitud:  {', '.join(header)}"
        )


def _reject_formulas(path: Path) -> None:
    """A formula in DATA means the file is not a deterministic export."""
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        sheet = workbook[DATA_SHEET]
        for row in sheet.iter_rows(min_row=2, max_col=len(DATA_COLUMNS)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    raise WorkbookContractError(
                        f"DATA lahtris {cell.coordinate} on valem. "
                        "Kanoonilises tabelis peavad olema arvutatud väärtused."
                    )
    finally:
        workbook.close()


def _parse_row(values: tuple, row_number: int) -> WorkbookRow:
    cells = dict(zip(DATA_COLUMNS, values))

    record_id = _text(cells["record_id"], limit=64)
    if not record_id:
        raise WorkbookContractError(f"Reas {row_number} puudub 'record_id'.")

    topic = _text(cells["topic"], limit=MAX_TOPIC_LENGTH)
    if not topic:
        raise WorkbookContractError(f"Reas {row_number} puudub 'topic'.")

    sent_status = _text(cells["sent_status"], limit=32)
    if sent_status not in VALID_SENT_STATUSES:
        raise WorkbookContractError(
            f"Reas {row_number} on tundmatu 'sent_status' {sent_status!r}. "
            f"Lubatud: {', '.join(sorted(VALID_SENT_STATUSES))}."
        )

    sent_date = _as_date(cells["sent_date"], column="sent_date", row_number=row_number)
    if sent_status == "sent" and sent_date is None:
        raise WorkbookContractError(f"Reas {row_number} on olek 'sent', kuid 'sent_date' puudub.")
    if sent_status != "sent" and sent_date is not None:
        raise WorkbookContractError(
            f"Reas {row_number} on 'sent_date' olemas, kuid olek on {sent_status!r}."
        )

    return WorkbookRow(
        record_id=record_id,
        source_year=_as_int(cells["source_year"], column="source_year", row_number=row_number),
        source_nr=_as_int(
            cells["source_nr"], column="source_nr", row_number=row_number, allow_none=True
        ),
        topic=topic,
        act_type=_text(cells["act_type"], limit=100),
        received_date=_as_date(
            cells["received_date"], column="received_date", row_number=row_number
        ),
        deadline_date=_as_date(
            cells["deadline_date"], column="deadline_date", row_number=row_number
        ),
        sent_date=sent_date,
        sent_status=sent_status,
        recipient=_text(cells["recipient"], limit=200),
        stage=_text(cells["stage"], limit=200),
        stage_key=_text(cells["stage_key"], limit=200),
        next_step=_text(cells["next_step"], limit=300),
        is_open=_as_bool(cells["is_open"], column="is_open", row_number=row_number),
        warning_codes=_split_warning_codes(cells["warning_codes"]),
        source_row=_as_int(cells["source_row"], column="source_row", row_number=row_number),
        refreshed_at=(
            cells["refreshed_at"] if isinstance(cells["refreshed_at"], dt.datetime) else None
        ),
    )


def _check_control_counts(control: WorkbookControl, parsed: ParsedWorkbook) -> None:
    """CONTROL must agree with what DATA actually contains.

    A workbook whose own summary disagrees with its rows is not trustworthy, so
    the mismatch fails the import rather than being silently overwritten with
    the recomputed value.
    """
    checks = (
        ("total_record_count", control.total_record_count, len(parsed.rows)),
        ("open_record_count", control.open_record_count, parsed.open_count),
        ("sent_record_count", control.sent_record_count, parsed.sent_count),
        ("not_sent_record_count", control.not_sent_record_count, parsed.not_sent_count),
        ("warning_record_count", control.warning_record_count, parsed.warning_record_count),
    )
    mismatches = [
        f"{name}: CONTROL {declared}, DATA {actual}"
        for name, declared, actual in checks
        if declared != actual
    ]
    if mismatches:
        raise WorkbookContractError(
            "CONTROL ei ole DATA lehega kooskõlas: " + "; ".join(mismatches) + "."
        )


def parse_workbook(path: Path | str) -> ParsedWorkbook:
    """Read and fully validate the workbook at ``path``."""
    path = Path(path)
    _reject_formulas(path)

    workbook = load_workbook(path, data_only=True)
    try:
        missing_sheets = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
        if missing_sheets:
            raise WorkbookContractError(f"Töövihikus puuduvad lehed: {', '.join(missing_sheets)}.")

        control = _require_control(_read_control(workbook[CONTROL_SHEET]))

        data_sheet = workbook[DATA_SHEET]
        _require_data_table(data_sheet)
        _require_header(data_sheet)

        rows: list[WorkbookRow] = []
        seen_records: set[str] = set()
        seen_source_rows: set[tuple[int, int]] = set()
        for offset, values in enumerate(
            data_sheet.iter_rows(min_row=2, max_col=len(DATA_COLUMNS), values_only=True),
            start=2,
        ):
            if all(value is None for value in values):
                continue
            row = _parse_row(values, offset)
            if row.record_id in seen_records:
                raise WorkbookContractError(f"Korduv 'record_id' {row.record_id!r} real {offset}.")
            seen_records.add(row.record_id)

            source_key = (row.source_year, row.source_row)
            if source_key in seen_source_rows:
                raise WorkbookContractError(
                    f"Korduv aasta ja lähterea kombinatsioon {source_key} real {offset}."
                )
            seen_source_rows.add(source_key)
            rows.append(row)

        warnings = tuple(_read_warnings(workbook[WARNINGS_SHEET]))
        parsed = ParsedWorkbook(control=control, rows=tuple(rows), warnings=warnings)
        _check_control_counts(control, parsed)
        return parsed
    finally:
        workbook.close()


def _read_warnings(sheet) -> list[dict]:
    """Structured warning rows, without the original offending values.

    ``original_value`` is deliberately dropped: it is source content, and the
    audit trail and import diagnostics must never accumulate workbook data.
    """
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    wanted = {"record_id", "source_row", "field", "warning_code"}
    indexes = {name: index for index, name in enumerate(header) if name in wanted}
    collected: list[dict] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in values):
            continue
        collected.append(
            {
                name: (_text(values[index], limit=64) if index < len(values) else "")
                for name, index in indexes.items()
            }
        )
    return collected
