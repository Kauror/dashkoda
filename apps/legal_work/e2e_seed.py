"""Synthetic legal-work content, published through the real workbook parser.

A generated XLSX goes through the same importer the OneDrive feed uses, so the
seeded state is one the application could actually have reached.
"""

from __future__ import annotations

import datetime as dt
import hashlib
from pathlib import Path

from django.core.management.base import CommandError

from apps.core.e2e_seed import LONG_TOPIC, freeze_package_timestamps

# Enough history that the annual series has a shape and the year-on-year
# comparison has a real baseline, without making the seeded import slow.
HISTORY_YEARS = 5

# Synthetic stages. The last two exercise the two cases the stage chart has to
# survive: a vocabulary entry nobody has seen before, and no stage at all.
SEED_STAGES: tuple[str, ...] = (
    "kooskõlastusringil",
    "Riigikogus",
    "ELi menetluses",
    "Eesti seisukoht",
    "sünteetiline uus etapp",
)

SEED_RECIPIENTS: tuple[str, ...] = (
    "Sünteetiline ministeerium A",
    "Sünteetiline ministeerium B",
    "Sünteetiline ministeerium C",
    "SÜNT-A",
)

SEED_ACT_TYPES: tuple[str, ...] = ("seadus", "määrus", "direktiiv", "VTK")


def _legal_work_rows(today: dt.date) -> list[list]:
    """DATA rows in canonical 1.2 column order, deliberately varied.

    Built to exercise the intelligence dashboard rather than just to fill the
    page, so the seeded browser run draws every shape the analytics can
    produce:

    - several complete historical years plus a partial current one, so the
      annual series and the `YTD` marking both have something to show;
    - sends on both sides of last year's same-date cutoff, so the year-on-year
      comparison is a real comparison rather than a part-year against a whole;
    - arrivals in every month, so the monthly charts are not one bar;
    - short, typical and long response windows, plus one deadline *before* its
      arrival, which must be excluded from the statistics and counted instead;
    - an arrival dated after the reporting date, which must never produce a
      negative age;
    - a matter whose deadline has passed with its opinion still pending, and
      one whose deadline has passed after its opinion went out — the two cases
      the deadline block must keep apart;
    - feedback counts that are absent, a measured zero, and positive, so the
      null-versus-zero distinction is visible on the page and not only in a
      unit test.
    """
    reporting_date = today - dt.timedelta(days=1)
    current_year = reporting_date.year
    rows: list[list] = []
    # `source_row` repeats across years by design, so it is numbered per year.
    per_year: dict[int, int] = {}

    def row(
        *,
        topic: str,
        source_year: int,
        received_date: dt.date | None,
        deadline_date: dt.date | None,
        sent_date: dt.date | None = None,
        sent_status: str = "pending",
        is_open: bool = True,
        warning_codes: str | None = None,
        stage: str = "kooskõlastusringil",
        recipient: str = "Sünteetiline ministeerium A",
        act_type: str = "seadus",
        feedback: int | None = None,
        requested: int | None = None,
    ) -> list:
        per_year[source_year] = per_year.get(source_year, 1) + 1
        source_row = per_year[source_year]
        return [
            f"SEED-{source_year}-{source_row:04d}",
            source_year,
            source_row,
            topic,
            act_type,
            received_date,
            deadline_date,
            sent_date,
            sent_status,
            recipient,
            stage,
            stage.lower(),
            "sünteetiline järgmine samm",
            is_open,
            warning_codes,
            source_row,
            dt.datetime.combine(reporting_date, dt.time(6, 30)),
            feedback,
            requested,
        ]

    def clamp(year: int, month: int, day: int) -> dt.date:
        """A safe date inside `month`, whatever its length."""
        return dt.date(year, month, min(day, 28))

    # ---------------------------------------------------------------- history
    # Complete years: a steady stream of concluded matters, so the annual chart
    # has finished bars to compare the current partial one against.
    for offset in range(HISTORY_YEARS, 0, -1):
        year = current_year - offset
        for month in range(1, 13, 2):
            received = clamp(year, month, 4)
            window = 7 + (month % 4) * 7
            rows.append(
                row(
                    topic=f"Sünteetiline {year}. aasta arvamus {month}",
                    source_year=year,
                    received_date=received,
                    deadline_date=received + dt.timedelta(days=window),
                    sent_date=received + dt.timedelta(days=window - 1),
                    sent_status="sent",
                    is_open=False,
                    stage="jõustunud",
                    recipient=SEED_RECIPIENTS[month % len(SEED_RECIPIENTS)],
                    act_type=SEED_ACT_TYPES[month % len(SEED_ACT_TYPES)],
                    # Feedback tracking starts partway through the history, so
                    # the coverage table has a real beginning rather than a
                    # decade of zeroes.
                    feedback=(month if offset <= 2 else None),
                    requested=(month * 6 if offset <= 2 else None),
                )
            )

    # -------------------------------------------------- the comparison year
    # Last year, with sends deliberately either side of the same-date cutoff.
    # Everything up to the cutoff is the baseline the current year is measured
    # against; everything after it belongs to the annual bar and to no
    # year-to-date figure.
    previous_year = current_year - 1
    for month in range(1, 13):
        received = clamp(previous_year, month, 6)
        rows.append(
            row(
                topic=f"Sünteetiline eelmise aasta teema {month}",
                source_year=previous_year,
                received_date=received,
                deadline_date=received + dt.timedelta(days=14),
                sent_date=received + dt.timedelta(days=12),
                sent_status="sent",
                is_open=False,
                stage="jõustunud",
                recipient=SEED_RECIPIENTS[month % len(SEED_RECIPIENTS)],
                act_type=SEED_ACT_TYPES[month % len(SEED_ACT_TYPES)],
                feedback=(0 if month % 4 == 0 else month),
                requested=month * 5,
            )
        )

    # ------------------------------------------------------- the current year
    # Arrivals in every month the year has reached, so the monthly chart is a
    # series rather than a single bar, and sends in most of them.
    for month in range(1, reporting_date.month + 1):
        received = clamp(current_year, month, 5)
        sent = received + dt.timedelta(days=11)
        concluded = sent <= reporting_date and month % 3 != 0
        rows.append(
            row(
                topic=f"Sünteetiline {current_year}. aasta teema {month}",
                source_year=current_year,
                received_date=received,
                deadline_date=received + dt.timedelta(days=13),
                sent_date=sent if concluded else None,
                sent_status="sent" if concluded else "pending",
                is_open=not concluded,
                stage=SEED_STAGES[month % len(SEED_STAGES)],
                recipient=SEED_RECIPIENTS[month % len(SEED_RECIPIENTS)],
                act_type=SEED_ACT_TYPES[month % len(SEED_ACT_TYPES)],
                feedback=(None if month % 5 == 0 else (0 if month % 3 == 0 else month * 2)),
                requested=(None if month % 5 == 0 else month * 9),
            )
        )

    # --------------------------------------------------- deliberate edge cases
    # The overflow candidate: an extremely long topic inside a linked cell.
    rows.append(
        row(
            topic=LONG_TOPIC,
            source_year=current_year,
            received_date=reporting_date - dt.timedelta(days=40),
            deadline_date=reporting_date + dt.timedelta(days=2),
            feedback=3,
            requested=30,
        )
    )
    # Deadlines across the urgency thresholds the selectors describe.
    for label, ahead in (("kiireloomuline", 1), ("lähituleviku", 8), ("rahulik", 18)):
        rows.append(
            row(
                topic=f"Sünteetiline {label} teema",
                source_year=current_year,
                received_date=reporting_date - dt.timedelta(days=30),
                deadline_date=reporting_date + dt.timedelta(days=ahead),
                stage=SEED_STAGES[ahead % len(SEED_STAGES)],
            )
        )
    # Open with no deadline at all: must sort last, never first.
    rows.append(
        row(
            topic="Sünteetiline tähtajata teema",
            source_year=current_year,
            received_date=reporting_date - dt.timedelta(days=15),
            deadline_date=None,
        )
    )
    # A warning code and an empty stage, which is what `Määramata` is drawn from.
    rows.append(
        row(
            topic="Sünteetiline hoiatusega teema",
            source_year=current_year,
            received_date=reporting_date - dt.timedelta(days=12),
            deadline_date=reporting_date + dt.timedelta(days=5),
            warning_codes="missing_stage",
            stage="",
        )
    )
    # Deadline passed with the opinion still pending: outstanding work.
    rows.append(
        row(
            topic="Sünteetiline möödunud tähtajaga ootel teema",
            source_year=current_year,
            received_date=reporting_date - dt.timedelta(days=60),
            deadline_date=reporting_date - dt.timedelta(days=10),
            stage="Riigikogus",
        )
    )
    # Deadline passed *after* the opinion went out, and the matter is still
    # open. Not late, and the page must not say it is.
    rows.append(
        row(
            topic="Sünteetiline möödunud tähtajaga saadetud teema",
            source_year=current_year,
            received_date=reporting_date - dt.timedelta(days=70),
            deadline_date=reporting_date - dt.timedelta(days=12),
            sent_date=reporting_date - dt.timedelta(days=20),
            sent_status="sent",
            is_open=True,
            stage="ootan jõustumist",
            feedback=0,
            requested=12,
        )
    )
    # An old open matter, so the age bands are not all one bucket.
    rows.append(
        row(
            topic="Sünteetiline pikaajaline ELi teema",
            source_year=current_year - 3,
            received_date=reporting_date - dt.timedelta(days=800),
            deadline_date=None,
            stage="ELi menetluses",
        )
    )
    # A deadline before its own arrival: a source-quality problem that must be
    # excluded from the response-window statistics and counted, never repaired.
    rows.append(
        row(
            topic="Sünteetiline vigase ajavahemikuga teema",
            source_year=current_year,
            received_date=reporting_date - dt.timedelta(days=20),
            deadline_date=reporting_date - dt.timedelta(days=30),
            warning_codes="deadline_before_received",
            stage="muu",
        )
    )
    # An arrival after the reporting date: must never produce a negative age.
    rows.append(
        row(
            topic="Sünteetiline tulevikku dateeritud teema",
            source_year=current_year,
            received_date=reporting_date + dt.timedelta(days=30),
            deadline_date=None,
            warning_codes="received_date_in_future",
            stage="idee",
        )
    )
    # No arrival date at all: belongs to the annual total and to no month.
    rows.append(
        row(
            topic="Sünteetiline kuupäevata teema",
            source_year=current_year,
            received_date=None,
            deadline_date=None,
            warning_codes="missing_received_date",
            stage="idee",
        )
    )
    # Explicitly not sent: distinct from pending, and carries no date.
    rows.append(
        row(
            topic="Sünteetiline saatmata jäetud teema",
            source_year=current_year,
            received_date=reporting_date - dt.timedelta(days=70),
            deadline_date=None,
            sent_status="not_sent",
            is_open=False,
            stage="rohkem tegevusi pole plaanis",
        )
    )
    # Enough remaining open rows that the bounded list actually scrolls.
    for index in range(1, 12):
        rows.append(
            row(
                topic=f"Sünteetiline töös olev teema {index}",
                source_year=current_year,
                received_date=reporting_date - dt.timedelta(days=index),
                deadline_date=reporting_date + dt.timedelta(days=index + 20),
                stage=SEED_STAGES[index % len(SEED_STAGES)],
                recipient=SEED_RECIPIENTS[index % len(SEED_RECIPIENTS)],
                act_type=SEED_ACT_TYPES[index % len(SEED_ACT_TYPES)],
                feedback=(None if index % 4 == 0 else index),
                requested=(None if index % 4 == 0 else index * 7),
            )
        )
    return rows


def write_workbook(path: Path, today: dt.date) -> Path:
    """Write a workbook that satisfies the canonical contract exactly."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    from apps.legal_work.workbook import (
        CONTROL_SHEET,
        DATA_COLUMNS,
        DATA_COLUMNS_V12,
        DATA_SHEET,
        DATA_TABLE_NAME,
        OVERVIEW_SHEET,
        WARNINGS_SHEET,
    )

    rows = _legal_work_rows(today)
    is_open_index = DATA_COLUMNS.index("is_open")
    sent_status_index = DATA_COLUMNS.index("sent_status")
    warnings_index = DATA_COLUMNS.index("warning_codes")
    # Read by name rather than from the end of the row: the 1.2 shape appends
    # two columns, so a negative index would now reach a feedback count.
    source_row_index = DATA_COLUMNS.index("source_row")
    reporting_date = today - dt.timedelta(days=1)
    generated_at = dt.datetime.combine(reporting_date, dt.time(6, 30))

    control = {
        "dataset_key": "oigusloome",
        # The seeded register declares 1.2, so the browser run exercises the
        # member-feedback analytics rather than only the base columns.
        "schema_version": "1.2",
        "source_file_name": "sünteetiline-lähtefail.xlsx",
        "source_sheet": "2099",
        "source_modified_at": "",
        "source_sha256": "0" * 64,
        "generated_at": generated_at.isoformat(),
        "reporting_date": reporting_date,
        "total_record_count": len(rows),
        "open_record_count": sum(1 for row in rows if row[is_open_index] is True),
        "sent_record_count": sum(1 for row in rows if row[sent_status_index] == "sent"),
        "not_sent_record_count": sum(1 for row in rows if row[sent_status_index] == "not_sent"),
        "warning_record_count": sum(1 for row in rows if row[warnings_index]),
        "refresh_status": "completed_with_warnings",
        "generator_version": "seed-e2e",
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in (CONTROL_SHEET, OVERVIEW_SHEET, DATA_SHEET, WARNINGS_SHEET):
        workbook.create_sheet(name)

    control_sheet = workbook[CONTROL_SHEET]
    control_sheet.append(["DASHKODA ÕIGUSLOOME ANDMEVOOG (SÜNTEETILINE)", None])
    for key, value in control.items():
        control_sheet.append([key, value])

    workbook[OVERVIEW_SHEET].append(["Sünteetiline ülevaade"])

    data = workbook[DATA_SHEET]
    data.append(list(DATA_COLUMNS_V12))
    for row in rows:
        data.append(row)
    last_column = chr(ord("A") + len(DATA_COLUMNS_V12) - 1)
    data.add_table(Table(displayName=DATA_TABLE_NAME, ref=f"A1:{last_column}{len(rows) + 1}"))

    warnings_sheet = workbook[WARNINGS_SHEET]
    warnings_sheet.append(["record_id", "source_row", "field", "warning_code"])
    for row in rows:
        if not row[warnings_index]:
            continue
        for code in str(row[warnings_index]).split(";"):
            warnings_sheet.append([row[0], row[source_row_index], "stage", code.strip()])
    warnings_sheet.add_table(
        Table(displayName="tbl_oigusloome_warnings", ref=f"A1:D{warnings_sheet.max_row}")
    )

    workbook.save(path)
    freeze_package_timestamps(path)
    return path


def seed(today: dt.date) -> str:
    from apps.legal_work.public_download import XLSX_MIME_TYPE, PublicDownload
    from apps.legal_work.public_sync import synchronize_public_workbook
    from apps.legal_work.sync import SyncLocked, advisory_lock

    def downloader(destination: Path) -> PublicDownload:
        write_workbook(destination, today)
        payload = destination.read_bytes()
        return PublicDownload(
            path=destination,
            size_bytes=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            content_type=XLSX_MIME_TYPE,
            final_host="synthetic-seed.invalid",
        )

    try:
        with advisory_lock():
            outcome = synchronize_public_workbook(downloader=downloader)
    except SyncLocked as error:
        raise CommandError(f"Legal-work seed could not take the lock: {error}") from error
    return f"õigusloome: {outcome.result} ({outcome.rows_imported} kirjet)"
