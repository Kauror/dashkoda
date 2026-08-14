"""Synthetic event programme, published through the real workbook importer.

The programme is the authority on every event figure the dashboard shows, so its
seed goes through the canonical parser rather than writing rows directly.
"""

from __future__ import annotations

import datetime as dt
import hashlib
from pathlib import Path

from django.core.management.base import CommandError

from apps.core.e2e_seed import freeze_package_timestamps

# The one event page the shop seed sells a registration product for. Imported
# rather than repeated, exactly as `apps.visibility.e2e_seed` imports it: three
# seeders have to name the same path or the cross-domain joins test nothing.
from apps.shop.e2e_seed import SHOP_EVENT_INDEX

# A synthetic public page on the allowed host. `/et/sundmused/` is the real
# prefix, and `sunteetiline-*` is unmistakably not a production path, so a
# screenshot can never be read as a link to a real Chamber event.
SYNTHETIC_EVENT_URL = "https://www.koda.ee/et/sundmused/sunteetiline-programmi-sundmus"


def _event_page(index: int) -> str:
    """A public page `apps.visibility.e2e_seed` has measured.

    The three seeders have to agree on these paths or the whole cross-domain
    half of this dashboard tests nothing: GA4 files traffic under a path, the
    shop files its event product under a path, and the programme's own
    `public_url` is what joins an event to both. `sunteetiline-3` in particular
    is `apps.shop.e2e_seed.SHOP_EVENT_INDEX`, which is the one event with
    Commerce registration facts behind it.
    """
    return f"https://www.koda.ee/et/sundmused/sunteetiline-{index}"


#: A page the programme links to that GA4 has **never measured**. Not a
#: zero-traffic page — an unmeasured one, which the interface must render as `—`
#: rather than as `0 vaatamist`.
UNMEASURED_EVENT_URL = "https://www.koda.ee/et/sundmused/sunteetiline-moootmata"

LONG_EVENT_NAME = (
    "Sünteetiline väga pikk sündmuse nimi, mis on kirjutatud ainult selleks, et "
    "kontrollida tabeliveeru kärpimist ja seda, kas pikk seotud pealkiri koos "
    "peidetud lisamärkusega ajab lehe horisontaalselt kerima; see ei kirjelda "
    "ühtegi tegelikku Koja sündmust"
)

# Estonian month names, as the Chamber's generator writes `event_month_label`.
EVENT_MONTH_LABELS = (
    "jaanuar",
    "veebruar",
    "märts",
    "aprill",
    "mai",
    "juuni",
    "juuli",
    "august",
    "september",
    "oktoober",
    "november",
    "detsember",
)

# Deliberately several of each, so tag and type filtering has something to
# separate and the option lists are not single-valued.
EVENT_TAGS = (
    ("seminar", "Sünteetiline seminar"),
    ("konverents", "Sünteetiline konverents"),
    ("koolitus", "Sünteetiline koolitus"),
)
EVENT_TYPES = (
    ("training", "Sünteetiline koolitusvorm"),
    ("conference", "Sünteetiline konverentsivorm"),
)
DELIVERY_MODES = ("onsite", "online", "hybrid")


def _programme_row(
    *,
    index: int,
    name: str,
    start: dt.date | None,
    end: dt.date | None = None,
    status: str,
    tag: tuple[str, str],
    event_type: tuple[str, str],
    delivery_mode: str,
    public_url: str | None = None,
    review_required: bool = False,
    price_status: str = "paid",
    member_price: float | None = 100,
    nonmember_price: float | None = 200,
    added_date: dt.date | None = None,
) -> dict:
    """One synthetic `DASH_EVENTS` row, keyed by the contract's column names.

    The four derived calendar fields come from `start` exactly as the generator
    derives them, so an undated row carries no year, month or quarter rather than
    an invented one.

    `planning_lead_days` is **derived here from the two dates**, exactly as the
    real generator derives it. Writing an independent number would make the seed
    the one place in the system where the source's own identity does not hold,
    and `tests/event_programme/test_planning.py` asserts that identity against
    imported rows.
    """
    tag_key, tag_label = tag
    type_key, type_label = event_type
    dated = start is not None
    lead = (start - added_date).days if dated and added_date is not None else None
    return {
        "event_id": f"SEED-EVENT-{index:04d}",
        "service_code": f"S{index:04d}",
        "event_name": name,
        "event_name_raw": name,
        "date_text_raw": start.isoformat() if dated else "sünteetiline kuupäevatekst",
        "start_date": dt.datetime.combine(start, dt.time()) if dated else None,
        "end_date": dt.datetime.combine(end or start, dt.time()) if dated else None,
        "event_year": start.year if dated else None,
        "event_month": start.month if dated else None,
        "event_month_key": f"{start:%Y-%m}" if dated else None,
        "event_month_label": EVENT_MONTH_LABELS[start.month - 1] if dated else None,
        "event_quarter": f"Q{(start.month - 1) // 3 + 1}" if dated else None,
        "event_status": status,
        "short_name_raw": "SÜN",
        "short_name_normalized": "syn",
        "tag_key": tag_key,
        "tag_label": tag_label,
        "event_type_key": type_key,
        "event_type_label": type_label,
        "delivery_mode": delivery_mode,
        "include_status": "REVIEW" if review_required else "YES",
        # Internal group columns exist in the export and must never reach a
        # model field: their business meaning has never been established.
        "group_raw": None,
        "group_secondary_raw": None,
        # The normalised price pair is stored; the `*_raw` echoes beside it are
        # not, and the later-price pair is not either.
        "member_price_raw": None if member_price is None else str(member_price),
        "member_price_eur": member_price,
        "nonmember_price_raw": None if nonmember_price is None else str(nonmember_price),
        "nonmember_price_eur": nonmember_price,
        "later_member_price_raw": None,
        "later_member_price_eur": None,
        "later_nonmember_price_raw": None,
        "later_nonmember_price_eur": None,
        "price_status": price_status,
        "discount_code": None,
        "discount_raw": None,
        "added_date": dt.datetime.combine(added_date, dt.time()) if added_date else None,
        "planning_lead_days": lead,
        "public_url": public_url,
        "public_link_status": "linked_embedded_latest" if public_url else "not_linked",
        "source_year": start.year if dated else 2099,
        # Deliberately not the event's own year: the page must never present the
        # annual sheet a row sat on as the event's date.
        "source_sheet": "KOOD 2099",
        "source_row": index + 1,
        "source_occurrence_count": 1,
        "date_parse_status": "parsed_range"
        if dated and end and end != start
        else ("parsed_single" if dated else "unparsed"),
        "review_required": review_required,
        "warning_codes": "date_unparsed" if not dated else None,
        "export_refreshed_at": None,
    }


def _programme_rows(today: dt.date) -> list[dict]:
    """A programme spanning several years, with every awkward shape in it.

    Covers past, ongoing and upcoming events, a multi-day range, a quarter
    boundary, an undated event, three tags, two types, all three delivery modes, a
    linked and an unlinked title, a review-required record, a very long linked
    name and enough rows to need more than one page.
    """
    rows: list[dict] = []
    index = 0

    def add(**kwargs) -> None:
        nonlocal index
        index += 1
        rows.append(_programme_row(index=index, **kwargs))

    def tag(position: int) -> tuple[str, str]:
        return EVENT_TAGS[position % len(EVENT_TAGS)]

    def event_type(position: int) -> tuple[str, str]:
        return EVENT_TYPES[position % len(EVENT_TYPES)]

    # The overflow candidate: a very long linked name in a table cell.
    add(
        name=LONG_EVENT_NAME,
        start=today + dt.timedelta(days=9),
        status="upcoming",
        tag=tag(0),
        event_type=event_type(0),
        delivery_mode="onsite",
        public_url=_event_page(1),
        added_date=today - dt.timedelta(days=51),
    )
    # Under way right now: started before today and ends after it.
    add(
        name="Sünteetiline käimasolev sündmus",
        start=today - dt.timedelta(days=1),
        end=today + dt.timedelta(days=1),
        status="ongoing",
        tag=tag(1),
        event_type=event_type(1),
        delivery_mode="hybrid",
        added_date=today - dt.timedelta(days=8),
    )
    # Inside the backward 30-day window, measured, and old enough that its whole
    # 30-day pre-event window sits inside GA4's coverage. The one event the
    # fair-comparison window can actually be computed for.
    add(
        name="Sünteetiline hiljuti toimunud sündmus",
        start=today - dt.timedelta(days=10),
        status="past",
        tag=tag(2),
        event_type=event_type(0),
        delivery_mode="online",
        public_url=_event_page(2),
        price_status="free",
        member_price=0,
        nonmember_price=0,
        added_date=today - dt.timedelta(days=100),
    )
    # The one event with Commerce registration facts behind it: its public page
    # is the shop seed's event-registration product page.
    add(
        name="Sünteetiline registreerimisega sündmus",
        start=today - dt.timedelta(days=12),
        status="past",
        tag=tag(0),
        event_type=event_type(1),
        delivery_mode="onsite",
        public_url=_event_page(SHOP_EVENT_INDEX),
        added_date=today - dt.timedelta(days=42),
    )
    # Linked to a page GA4 has never measured. Its view column must read `—`,
    # not `0`, and it must sort behind every measured event rather than as
    # though it had scored nothing.
    add(
        name="Sünteetiline mõõtmata lehega sündmus",
        start=today - dt.timedelta(days=14),
        status="past",
        tag=tag(1),
        event_type=event_type(0),
        delivery_mode="online",
        public_url=UNMEASURED_EVENT_URL,
        price_status="mixed",
        member_price=0,
        nonmember_price=45,
        added_date=today - dt.timedelta(days=70),
    )
    # Two programme events pointing at one public page. Their traffic may not be
    # counted twice, and neither may be silently dropped.
    for suffix, offset in (("A", 16), ("B", 17)):
        add(
            name=f"Sünteetiline jagatud lehega sündmus {suffix}",
            start=today - dt.timedelta(days=offset),
            status="past",
            tag=tag(2),
            event_type=event_type(1),
            delivery_mode="hybrid",
            public_url=_event_page(4),
            added_date=today - dt.timedelta(days=offset + 30),
        )
    # Delivery mode the source never stated. It is `Määramata`, never `Kohapeal`.
    add(
        name="Sünteetiline määramata toimumisviisiga sündmus",
        start=today - dt.timedelta(days=20),
        status="past",
        tag=tag(0),
        event_type=event_type(0),
        delivery_mode="",
        price_status="tba",
        member_price=None,
        nonmember_price=None,
        added_date=today - dt.timedelta(days=25),
    )
    # A price nobody recorded. It must never be presented as free.
    add(
        name="Sünteetiline teadmata hinnaga sündmus",
        start=today - dt.timedelta(days=22),
        status="past",
        tag=tag(1),
        event_type=event_type(1),
        delivery_mode="onsite",
        price_status="missing",
        member_price=None,
        nonmember_price=None,
        added_date=today - dt.timedelta(days=120),
    )
    # Entered into the programme *after* it began. A real data-quality fact:
    # kept, flagged, and excluded from the planning statistics rather than
    # clamped to zero.
    add(
        name="Sünteetiline tagantjärele lisatud sündmus",
        start=today - dt.timedelta(days=30),
        status="past",
        tag=tag(2),
        event_type=event_type(0),
        delivery_mode="onsite",
        added_date=today - dt.timedelta(days=5),
    )
    # No `added_date` at all: planning coverage is never 100%, and the focus has
    # to state its own denominator.
    add(
        name="Sünteetiline planeerimisandmeteta sündmus",
        start=today - dt.timedelta(days=35),
        status="past",
        tag=tag(0),
        event_type=event_type(1),
        delivery_mode="online",
        added_date=None,
    )
    # Inside the forward 30-day window, as a multi-day range, and receiving web
    # attention right now.
    add(
        name="Sünteetiline mitmepäevane sündmus",
        start=today + dt.timedelta(days=3),
        end=today + dt.timedelta(days=5),
        status="upcoming",
        tag=tag(0),
        event_type=event_type(1),
        delivery_mode="onsite",
        public_url=_event_page(5),
        added_date=today - dt.timedelta(days=97),
    )
    # An eight-day programme, so the duration bands have a long entry.
    add(
        name="Sünteetiline mitmenädalane programm",
        start=today + dt.timedelta(days=12),
        end=today + dt.timedelta(days=19),
        status="upcoming",
        tag=tag(1),
        event_type=event_type(0),
        delivery_mode="hybrid",
        public_url=_event_page(6),
        member_price=450,
        nonmember_price=900,
        added_date=today - dt.timedelta(days=200),
    )
    # Starts inside the next 30 days with no public page at all: the actionable
    # signal the overview surfaces under `Tähelepanu`.
    add(
        name="Sünteetiline avaliku leheta tulevane sündmus",
        start=today + dt.timedelta(days=6),
        status="upcoming",
        tag=tag(2),
        event_type=event_type(0),
        delivery_mode="online",
        added_date=today - dt.timedelta(days=11),
    )
    # The quarter boundary: 31 March and 1 April of the same year.
    boundary_year = today.year - 1
    add(
        name="Sünteetiline kvartali lõpu sündmus",
        start=dt.date(boundary_year, 3, 31),
        status="past",
        tag=tag(1),
        event_type=event_type(0),
        delivery_mode="online",
    )
    add(
        name="Sünteetiline kvartali alguse sündmus",
        start=dt.date(boundary_year, 4, 1),
        status="past",
        tag=tag(2),
        event_type=event_type(1),
        delivery_mode="hybrid",
    )
    # A real event whose operational row held date text nobody could parse. It
    # must stay a record, must not acquire a date, and must be reachable.
    add(
        name="Sünteetiline kuupäevata sündmus",
        start=None,
        status="date_unknown",
        tag=tag(0),
        event_type=event_type(0),
        delivery_mode="onsite",
        review_required=True,
    )
    # Enough history that year filtering has several years to choose between and
    # that the whole programme runs past one page of 50 rows.
    #
    # The price and planning variety is spread deterministically rather than
    # randomly: a seed that produces a different distribution on each run cannot
    # be asserted against.
    for offset in range(1, 4):
        year = today.year - offset
        for month in (2, 3, 5, 9, 11, 12):
            for day in (7, 14, 21):
                position = month + day
                free = position % 4 == 0
                add(
                    name=f"Sünteetiline sündmus {year}-{month:02d}-{day:02d}",
                    start=dt.date(year, month, day),
                    status="past",
                    tag=tag(position),
                    event_type=event_type(month),
                    delivery_mode=DELIVERY_MODES[position % len(DELIVERY_MODES)],
                    price_status="free" if free else "paid",
                    member_price=0 if free else 40 + position,
                    nonmember_price=0 if free else 80 + position * 2,
                    # Leads across every band the planning focus draws.
                    added_date=dt.date(year, month, day) - dt.timedelta(days=7 + position * 3),
                )
    return rows


def write_workbook(path: Path, today: dt.date) -> Path:
    """Write a workbook that satisfies the canonical event contract exactly.

    Built here rather than imported from the test suite: this command ships in the
    application image and may not depend on `tests/`. It writes the same six
    sheets, the same two-row banner and the same text CONTROL table the Chamber's
    Office Script produces, so the real parser is genuinely exercised.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    from apps.event_programme.workbook import (
        CONTROL_SHEET,
        CONTROL_TABLE_NAME,
        EVENTS_COLUMNS,
        EVENTS_SHEET,
        EVENTS_TABLE_NAME,
        OCCURRENCES_SHEET,
        REVIEW_SHEET,
        TAG_MAP_SHEET,
        URL_OVERRIDES_SHEET,
    )

    # The banner rows the generator writes above every DASH_* table.
    banner_rows = 2
    control_banner_rows = 4

    rows = _programme_rows(today)
    export_refreshed_at = dt.datetime.combine(today, dt.time(6, 30)).isoformat()
    for row in rows:
        row["export_refreshed_at"] = export_refreshed_at

    linked = sum(1 for row in rows if row["public_url"])
    control = {
        "dataset_key": "events",
        "schema_version": "1.0",
        "generator_name": "Sünteetiline seemnegeneraator",
        "generator_version": "seed-e2e",
        "source_workbook_name": "sünteetiline-lähtefail.xlsx",
        "refresh_status": "ready_with_warnings",
        "canonical_event_count": str(len(rows)),
        "qualifying_occurrence_count": str(len(rows)),
        "excluded_event_count": "0",
        "repeated_service_code_count": "0",
        "linked_public_url_count": str(linked),
        "distinct_short_name_count": "1",
        "blocking_error_count": "0",
        "warning_count": str(sum(1 for row in rows if row["warning_codes"])),
        "export_refreshed_at": export_refreshed_at,
        "last_successful_refresh_at": export_refreshed_at,
    }

    def write_table(sheet, *, table_name, columns, table_rows, banners) -> None:
        for offset in range(banners):
            sheet.cell(row=offset + 1, column=1, value=f"Sünteetiline bänner {offset + 1}")
        header_row = banners + 1
        for position, name in enumerate(columns, start=1):
            sheet.cell(row=header_row, column=position, value=name)
        for row_offset, values in enumerate(table_rows, start=header_row + 1):
            for position, value in enumerate(values, start=1):
                sheet.cell(row=row_offset, column=position, value=value)
        last_row = header_row + max(len(table_rows), 1)
        last_column = sheet.cell(row=header_row, column=len(columns)).column_letter
        sheet.add_table(Table(displayName=table_name, ref=f"A{header_row}:{last_column}{last_row}"))

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_table(
        workbook.create_sheet(CONTROL_SHEET),
        table_name=CONTROL_TABLE_NAME,
        columns=("key", "value"),
        table_rows=[(key, value) for key, value in control.items()],
        banners=control_banner_rows,
    )
    write_table(
        workbook.create_sheet(EVENTS_SHEET),
        table_name=EVENTS_TABLE_NAME,
        columns=EVENTS_COLUMNS,
        table_rows=[tuple(row.get(name) for name in EVENTS_COLUMNS) for row in rows],
        banners=banner_rows,
    )
    # Present with their tables, never read as a data source.
    for sheet_name, table_name, columns in (
        (OCCURRENCES_SHEET, "tbl_dash_event_occurrences", ("event_id", "occurrence_number")),
        (REVIEW_SHEET, "tbl_dash_review", ("severity", "issue_code")),
        (TAG_MAP_SHEET, "tbl_dash_tag_map", ("short_name_normalized", "tag_key")),
        (URL_OVERRIDES_SHEET, "tbl_dash_url_overrides", ("service_code", "public_url")),
    ):
        write_table(
            workbook.create_sheet(sheet_name),
            table_name=table_name,
            columns=columns,
            table_rows=[],
            banners=banner_rows,
        )

    workbook.save(path)
    freeze_package_timestamps(path)
    return path


def seed(today: dt.date) -> str:
    """Publish the programme through the real synchronisation path.

    The synthetic workbook is handed to the same `synchronize_public_workbook`
    production uses, so the parser, the contract validation, the import registry,
    the atomic publication and the audit trail are all exercised. No
    `EventProgrammeItem` row is written directly, and nothing here contacts
    OneDrive: the downloader writes the file the caller asked for.
    """
    from apps.event_programme.public_download import XLSX_MIME_TYPE, PublicDownload
    from apps.event_programme.sync import SyncLocked, advisory_lock, synchronize_public_workbook

    def downloader(destination: Path) -> PublicDownload:
        write_workbook(destination, today)
        payload = destination.read_bytes()
        return PublicDownload(
            path=destination,
            size_bytes=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            content_type=XLSX_MIME_TYPE,
            final_host="synthetic-seed.invalid",
        )

    try:
        with advisory_lock():
            outcome = synchronize_public_workbook(downloader=downloader)
    except SyncLocked as error:
        raise CommandError(f"Event-programme seed could not take the lock: {error}") from error
    return f"sündmuste programm: {outcome.result} ({outcome.rows_imported} sündmust)"
