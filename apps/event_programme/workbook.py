"""Deterministic reader for the canonical event-programme workbook.

This module knows the workbook contract and nothing about Django models, the
import registry or OneDrive. It either returns a fully validated
:class:`ParsedWorkbook` or raises :class:`WorkbookContractError`.

The authoritative input is the Excel Table ``tbl_dash_events`` on the
``DASH_EVENTS`` sheet. Four further sheets travel with the file and are
**checked for presence but never read as a data source**:

- ``DASH_EVENT_OCCURRENCES`` — the raw rows behind each canonical event;
- ``DASH_REVIEW`` — the generator's own quality findings;
- ``DASH_TAG_MAP`` and ``DASH_URL_OVERRIDES`` — hand-maintained inputs whose
  authoritative copy lives in the Chamber's operational workbook.

They are the generator's working material. Requiring them proves the file is a
complete export rather than a truncated one, while storing them would put
someone else's editing surface into DashKoda.

Unlike the legal-work export, every DASH_* table starts at **row 3**: the two
rows above it carry the workbook's own banner for the people who maintain it.
The header row is therefore located from the table's declared range rather than
assumed, so a future banner change cannot silently shift the parser onto the
wrong row.

Nothing here interprets the meaning of an event. Values are parsed, type-checked
and rejected — never repaired, inferred or summarised.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

DATASET_KEY = "events"

# The workbook's own declared contract version. Only 1.0 exists so far; the
# importer holds an explicit supported set rather than accepting anything, and
# the value a file declared is recorded on its snapshot.
SUPPORTED_SCHEMA_VERSIONS = ("1.0",)

CONTROL_SHEET = "DASH_CONTROL"
EVENTS_SHEET = "DASH_EVENTS"
OCCURRENCES_SHEET = "DASH_EVENT_OCCURRENCES"
REVIEW_SHEET = "DASH_REVIEW"
TAG_MAP_SHEET = "DASH_TAG_MAP"
URL_OVERRIDES_SHEET = "DASH_URL_OVERRIDES"

REQUIRED_SHEETS = (
    CONTROL_SHEET,
    EVENTS_SHEET,
    OCCURRENCES_SHEET,
    REVIEW_SHEET,
    TAG_MAP_SHEET,
    URL_OVERRIDES_SHEET,
)

CONTROL_TABLE_NAME = "tbl_dash_control"
EVENTS_TABLE_NAME = "tbl_dash_events"

# Sheet -> the Excel Table it must carry. A sheet present without its table is a
# broken export, not an empty one.
REQUIRED_TABLES = {
    CONTROL_SHEET: CONTROL_TABLE_NAME,
    EVENTS_SHEET: EVENTS_TABLE_NAME,
    OCCURRENCES_SHEET: "tbl_dash_event_occurrences",
    REVIEW_SHEET: "tbl_dash_review",
    TAG_MAP_SHEET: "tbl_dash_tag_map",
    URL_OVERRIDES_SHEET: "tbl_dash_url_overrides",
}

# Exact names, in exact order, as the generator writes them. A reordered or
# renamed column is a contract break, not something to be matched up by
# guesswork. DashKoda stores only the subset named in STORED_COLUMNS below; the
# rest are verified so that a generator change cannot pass unnoticed, then
# discarded without ever reaching a model field.
EVENTS_COLUMNS = (
    "event_id",
    "service_code",
    "event_name",
    "event_name_raw",
    "date_text_raw",
    "start_date",
    "end_date",
    "event_year",
    "event_month",
    "event_month_key",
    "event_month_label",
    "event_quarter",
    "event_status",
    "short_name_raw",
    "short_name_normalized",
    "tag_key",
    "tag_label",
    "event_type_key",
    "event_type_label",
    "delivery_mode",
    "include_status",
    "group_raw",
    "group_secondary_raw",
    "member_price_raw",
    "member_price_eur",
    "nonmember_price_raw",
    "nonmember_price_eur",
    "later_member_price_raw",
    "later_member_price_eur",
    "later_nonmember_price_raw",
    "later_nonmember_price_eur",
    "price_status",
    "discount_code",
    "discount_raw",
    "added_date",
    "planning_lead_days",
    "public_url",
    "public_link_status",
    "source_year",
    "source_sheet",
    "source_row",
    "source_occurrence_count",
    "date_parse_status",
    "review_required",
    "warning_codes",
    "export_refreshed_at",
)

# What DashKoda keeps.
#
# The ``*_raw`` echo columns, the discount pair and the internal group columns
# stay deliberately absent from the model as well as the interface: the product
# does not need them, their business meaning has never been established, and a
# field that does not exist cannot leak. They are parsed past, not parsed in.
#
# The five **normalised** planning and price columns are stored, and that is a
# deliberate reversal of the first implementation. They were profiled against
# the real export before the fields were added:
#
# - ``added_date`` on 1 183 of 1 190 rows, spread across every source year in
#   proportion to that year's event count — so it is a real per-event date
#   rather than a backfill stamped on the day somebody built the sheet;
# - ``planning_lead_days`` on 1 151 rows, and equal to
#   ``start_date - added_date`` on **every** row where both exist. The
#   generator's own arithmetic is therefore reproduced rather than replaced;
# - ``price_status`` on 100% of rows in every source year, which is what makes
#   "free" a stated fact instead of an absent price read as zero;
# - the two current price columns, on 96% and 94% of rows.
#
# ``later_member_price_eur`` and ``later_nonmember_price_eur`` are still
# discarded: 3.4% coverage and no documented meaning for "later" is not enough
# to put a number on a screen.
STORED_COLUMNS = (
    "event_id",
    "service_code",
    "event_name",
    "start_date",
    "end_date",
    "event_year",
    "event_month_key",
    "event_month_label",
    "event_quarter",
    "event_status",
    "tag_key",
    "tag_label",
    "event_type_key",
    "event_type_label",
    "delivery_mode",
    "include_status",
    "member_price_eur",
    "nonmember_price_eur",
    "price_status",
    "added_date",
    "planning_lead_days",
    "public_url",
    "public_link_status",
    "source_year",
    "source_sheet",
    "source_row",
    "source_occurrence_count",
    "date_parse_status",
    "review_required",
    "warning_codes",
)

REQUIRED_CONTROL_KEYS = (
    "dataset_key",
    "schema_version",
    "generator_name",
    "generator_version",
    "source_workbook_name",
    "refresh_status",
    "canonical_event_count",
    "qualifying_occurrence_count",
    "excluded_event_count",
    "repeated_service_code_count",
    "linked_public_url_count",
    "distinct_short_name_count",
    "blocking_error_count",
    "warning_count",
    "export_refreshed_at",
    "last_successful_refresh_at",
)

CONTROL_COUNT_KEYS = (
    "canonical_event_count",
    "qualifying_occurrence_count",
    "excluded_event_count",
    "repeated_service_code_count",
    "linked_public_url_count",
    "distinct_short_name_count",
    "blocking_error_count",
    "warning_count",
)

# The generator publishes ``ready`` when nothing needed a human, and
# ``ready_with_warnings`` when the export is complete but some rows carry
# advisory codes. Anything else — ``blocked``, a partial refresh, a status this
# importer has never heard of — is not importable.
IMPORTABLE_REFRESH_STATUSES = frozenset({"ready", "ready_with_warnings"})

# A row is part of the published programme only when the generator says so.
INCLUDE_STATUS_YES = "YES"

# The generator separates codes with ";"; "," is tolerated defensively because
# both have appeared in hand-checked exports.
WARNING_CODE_SEPARATORS = (";", ",")

MAX_NAME_LENGTH = 500
MAX_SHORT_TEXT_LENGTH = 200
MAX_URL_LENGTH = 500


class WorkbookContractError(ValueError):
    """The workbook does not satisfy the agreed contract."""


@dataclass(frozen=True)
class WorkbookControl:
    dataset_key: str
    schema_version: str
    generator_name: str
    generator_version: str
    source_workbook_name: str
    refresh_status: str
    canonical_event_count: int
    qualifying_occurrence_count: int
    excluded_event_count: int
    repeated_service_code_count: int
    linked_public_url_count: int
    distinct_short_name_count: int
    blocking_error_count: int
    warning_count: int
    export_refreshed_at: dt.datetime
    last_successful_refresh_at: dt.datetime | None


@dataclass(frozen=True)
class WorkbookRow:
    event_id: str
    service_code: str
    event_name: str
    # Null when the generator could not parse a date from the operational
    # sheet's free text. That is "the source did not say", never "no date".
    start_date: dt.date | None
    end_date: dt.date | None
    event_year: int | None
    event_month_key: str
    event_month_label: str
    event_quarter: str
    event_status: str
    tag_key: str
    tag_label: str
    event_type_key: str
    event_type_label: str
    delivery_mode: str
    include_status: str
    # The event's current list prices and the generator's own statement about
    # them. `price_status` is authoritative: a null price with status `free` is
    # a free event, and a null price with status `missing`, `tba` or `review` is
    # an unknown one. The two are never collapsed.
    member_price_eur: Decimal | None
    nonmember_price_eur: Decimal | None
    price_status: str
    # When the event entered the operational programme, and the generator's own
    # `start_date - added_date`. Negative values are real and are kept: an event
    # entered after it ran is a data-entry fact, not something to clamp to zero.
    added_date: dt.date | None
    planning_lead_days: int | None
    public_url: str
    public_link_status: str
    source_year: int
    source_sheet: str
    source_row: int
    source_occurrence_count: int
    date_parse_status: str
    review_required: bool
    warning_codes: list[str]


@dataclass(frozen=True)
class ParsedWorkbook:
    control: WorkbookControl
    rows: tuple[WorkbookRow, ...]

    @property
    def linked_public_url_count(self) -> int:
        return sum(1 for row in self.rows if row.public_url)

    @property
    def review_required_count(self) -> int:
        return sum(1 for row in self.rows if row.review_required)

    @property
    def dated_event_count(self) -> int:
        """Events the generator could place on the calendar.

        The difference between this and the row count is not an error: an event
        whose operational row holds unparseable date text is still a real event
        that the Chamber ran.
        """
        return sum(1 for row in self.rows if row.start_date is not None)


def _text(value, *, limit: int = MAX_SHORT_TEXT_LENGTH) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:limit]


def _is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _as_date(value, *, column: str, row_number: int) -> dt.date | None:
    if _is_blank(value):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise WorkbookContractError(
        f"Reas {row_number} ei ole veerg {column!r} kuupäev vaid {type(value).__name__}."
    )


def _as_int(value, *, column: str, row_number: int, allow_none: bool = False) -> int | None:
    if _is_blank(value):
        if allow_none:
            return None
        raise WorkbookContractError(f"Reas {row_number} puudub kohustuslik veerg {column!r}.")
    if isinstance(value, bool) or not isinstance(value, int):
        raise WorkbookContractError(
            f"Reas {row_number} ei ole veerg {column!r} täisarv vaid {type(value).__name__}."
        )
    return value


def _as_money(value, *, column: str, row_number: int) -> Decimal | None:
    """A parsed euro amount, or None when the generator left the cell empty.

    Only the **normalised** price columns come through here, and they are
    numeric by contract — the generator has already read whatever the
    operational sheet said and either produced a number or left the cell blank
    and said so in ``price_status``. Text is therefore a contract break rather
    than something to interpret: a price this importer tried to read out of
    ``"al. 50€"`` would be a number nobody verified.

    Zero is a value, not a blank. It is what the generator writes for a free
    event, and `price_status` is what says so.
    """
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        raise WorkbookContractError(
            f"Reas {row_number} ei ole veerg {column!r} arv vaid tõeväärtus."
        )
    if not isinstance(value, (int, float, Decimal)):
        raise WorkbookContractError(
            f"Reas {row_number} ei ole veerg {column!r} arv vaid {type(value).__name__}."
        )
    try:
        # Via `str`, so 40.0 does not arrive as 40.000000000000001.
        amount = Decimal(str(value))
    except InvalidOperation as error:  # pragma: no cover - guarded above
        raise WorkbookContractError(
            f"Reas {row_number} ei ole veerg {column!r} loetav arv."
        ) from error
    if amount < 0:
        raise WorkbookContractError(f"Reas {row_number} on veerg {column!r} negatiivne.")
    return amount.quantize(Decimal("0.01"))


def _as_bool(value, *, column: str, row_number: int) -> bool:
    if isinstance(value, bool):
        return value
    raise WorkbookContractError(
        f"Reas {row_number} ei ole veerg {column!r} tõeväärtus vaid {type(value).__name__}."
    )


def _split_warning_codes(value) -> list[str]:
    if value is None:
        return []
    text = str(value)
    for separator in WARNING_CODE_SEPARATORS[1:]:
        text = text.replace(separator, WARNING_CODE_SEPARATORS[0])
    return [part.strip() for part in text.split(WARNING_CODE_SEPARATORS[0]) if part.strip()]


def _as_control_int(value, *, key: str) -> int:
    """Read a CONTROL count.

    The generator writes the whole CONTROL sheet as a text key/value table, so
    every count arrives as a string even though it is conceptually a number.
    That is a property of the Office Script rather than a defect, and this is
    the one place the importer accommodates it — by accepting a plain decimal
    string, not by coercing whatever turns up. ``"12.0"``, ``"1 186"``, ``TRUE``
    and an empty cell are all rejected.
    """
    if isinstance(value, bool):
        raise WorkbookContractError(f"DASH_CONTROL väli {key!r} peab olema täisarv.")
    if isinstance(value, int):
        number = value
    elif isinstance(value, str):
        text = value.strip()
        if not (text.isdigit() or (text.startswith("-") and text[1:].isdigit())):
            raise WorkbookContractError(f"DASH_CONTROL väli {key!r} peab olema täisarv.")
        number = int(text)
    else:
        raise WorkbookContractError(f"DASH_CONTROL väli {key!r} peab olema täisarv.")
    if number < 0:
        raise WorkbookContractError(f"DASH_CONTROL väli {key!r} ei tohi olla negatiivne.")
    return number


def _parse_timestamp(value, *, key: str) -> dt.datetime | None:
    if _is_blank(value):
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, str):
        try:
            return dt.datetime.fromisoformat(value.strip())
        except ValueError as error:
            raise WorkbookContractError(f"CONTROL väli {key!r} ei ole kehtiv ajatempel.") from error
    raise WorkbookContractError(f"CONTROL väli {key!r} ei ole ajatempel.")


def _header_row_number(sheet, table_name: str) -> int:
    """Locate a table's header row from its declared range.

    The DASH_* sheets carry a banner above the table, so the header is not on
    row 1 and its position is a property of the file rather than a constant to
    hard-code here.
    """
    tables = dict(getattr(sheet, "tables", {}) or {})
    if table_name not in tables:
        found = ", ".join(sorted(tables)) or "(ühtegi)"
        raise WorkbookContractError(
            f"Lehel {sheet.title!r} puudub Exceli tabel {table_name!r}. Leitud: {found}."
        )
    table = tables[table_name]
    reference = table if isinstance(table, str) else table.ref
    first_cell = reference.split(":")[0]
    digits = "".join(character for character in first_cell if character.isdigit())
    if not digits:
        raise WorkbookContractError(f"Tabeli {table_name!r} vahemik {reference!r} ei ole loetav.")
    return int(digits)


def _require_sheets(workbook) -> None:
    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise WorkbookContractError(f"Töövihikus puuduvad lehed: {', '.join(sorted(missing))}.")


def _require_tables(workbook) -> None:
    for sheet_name, table_name in REQUIRED_TABLES.items():
        _header_row_number(workbook[sheet_name], table_name)


def _read_control(sheet) -> dict:
    """Collect the key/value pairs from the CONTROL table.

    Rows above the table are the workbook's own banner and instructions for the
    people who maintain it, so reading starts at the table's header. A row that
    has a key keeps it even when the value cell is empty: "the generator did not
    write this field" and "the generator wrote it and it is legitimately empty"
    are different facts, and only :func:`_require_control` decides which keys may
    be empty. Unknown keys are ignored rather than rejected, so the generator can
    add metadata without breaking this importer.
    """
    header_row = _header_row_number(sheet, CONTROL_TABLE_NAME)
    values: dict = {}
    for key_cell, value_cell in sheet.iter_rows(
        min_row=header_row + 1, max_col=2, values_only=True
    ):
        if key_cell is None:
            continue
        key = str(key_cell).strip()
        if key:
            values[key] = value_cell
    return values


def _require_control(values: dict) -> WorkbookControl:
    missing = [key for key in REQUIRED_CONTROL_KEYS if key not in values]
    if missing:
        raise WorkbookContractError(
            f"DASH_CONTROL lehelt puuduvad väljad: {', '.join(sorted(missing))}."
        )

    blank = [
        key
        for key in REQUIRED_CONTROL_KEYS
        if key != "last_successful_refresh_at" and _is_blank(values[key])
    ]
    if blank:
        raise WorkbookContractError(
            f"DASH_CONTROL lehe väljad on tühjad: {', '.join(sorted(blank))}."
        )

    dataset_key = _text(values["dataset_key"])
    if dataset_key != DATASET_KEY:
        raise WorkbookContractError(
            f"Vale andmestik: oodati {DATASET_KEY!r}, sain {dataset_key!r}."
        )

    schema_version = _text(values["schema_version"])
    if schema_version not in SUPPORTED_SCHEMA_VERSIONS:
        raise WorkbookContractError(
            f"Toetamata skeemi versioon {schema_version!r}. "
            f"Toetatud: {', '.join(SUPPORTED_SCHEMA_VERSIONS)}."
        )

    refresh_status = _text(values["refresh_status"])
    if refresh_status not in IMPORTABLE_REFRESH_STATUSES:
        raise WorkbookContractError(
            f"Ekspordi seisund {refresh_status!r} ei ole imporditav. "
            f"Lubatud: {', '.join(sorted(IMPORTABLE_REFRESH_STATUSES))}."
        )

    counts = {key: _as_control_int(values[key], key=key) for key in CONTROL_COUNT_KEYS}

    # The generator's own gate. A file that admits to a blocking error is not
    # repaired here and is not imported: the fix belongs to whatever produced it.
    if counts["blocking_error_count"]:
        raise WorkbookContractError(
            f"Eksport teatab {counts['blocking_error_count']} blokeerivast veast; "
            "faili ei impordita."
        )

    export_refreshed_at = _parse_timestamp(values["export_refreshed_at"], key="export_refreshed_at")
    if export_refreshed_at is None:
        raise WorkbookContractError("DASH_CONTROL väli 'export_refreshed_at' on kohustuslik.")

    return WorkbookControl(
        dataset_key=dataset_key,
        schema_version=schema_version,
        generator_name=_text(values["generator_name"]),
        generator_version=_text(values["generator_version"]),
        source_workbook_name=_text(values["source_workbook_name"]),
        refresh_status=refresh_status,
        export_refreshed_at=export_refreshed_at,
        last_successful_refresh_at=_parse_timestamp(
            values["last_successful_refresh_at"], key="last_successful_refresh_at"
        ),
        **counts,
    )


def _require_header(sheet, header_row: int) -> None:
    header = tuple(
        _text(cell.value, limit=200) for cell in sheet[header_row][: len(EVENTS_COLUMNS)]
    )
    if header != EVENTS_COLUMNS:
        raise WorkbookContractError(
            f"{EVENTS_SHEET} lehe veerud ei vasta kokkuleppele.\n"
            f"Oodati: {', '.join(EVENTS_COLUMNS)}\n"
            f"Sain:   {', '.join(header)}"
        )


def _parse_rows(sheet, header_row: int) -> tuple[WorkbookRow, ...]:
    index = {name: position for position, name in enumerate(EVENTS_COLUMNS)}
    rows: list[WorkbookRow] = []
    seen_event_ids: set[str] = set()
    seen_service_codes: set[str] = set()

    for offset, raw in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        if raw is None or all(_is_blank(cell) for cell in raw[: len(EVENTS_COLUMNS)]):
            continue

        def cell(name: str):
            return raw[index[name]]

        event_id = _text(cell("event_id"))
        if not event_id:
            raise WorkbookContractError(f"Reas {offset} puudub 'event_id'.")
        if event_id in seen_event_ids:
            raise WorkbookContractError(f"'event_id' {event_id!r} kordub reas {offset}.")
        seen_event_ids.add(event_id)

        service_code = _text(cell("service_code"))
        if not service_code:
            raise WorkbookContractError(f"Reas {offset} puudub 'service_code'.")
        # The canonical table holds one row per service code; repeats belong to
        # DASH_EVENT_OCCURRENCES. A duplicate here means the generator collapsed
        # nothing, and the counts below would silently disagree.
        if service_code in seen_service_codes:
            raise WorkbookContractError(f"'service_code' {service_code!r} kordub reas {offset}.")
        seen_service_codes.add(service_code)

        event_name = _text(cell("event_name"), limit=MAX_NAME_LENGTH)
        if not event_name:
            raise WorkbookContractError(f"Reas {offset} puudub 'event_name'.")

        start_date = _as_date(cell("start_date"), column="start_date", row_number=offset)
        end_date = _as_date(cell("end_date"), column="end_date", row_number=offset)
        if start_date is not None and end_date is not None and end_date < start_date:
            raise WorkbookContractError(f"Reas {offset} on 'end_date' varasem kui 'start_date'.")
        if start_date is None and end_date is not None:
            raise WorkbookContractError(
                f"Reas {offset} on 'end_date' ilma 'start_date' väärtuseta."
            )

        rows.append(
            WorkbookRow(
                event_id=event_id,
                service_code=service_code,
                event_name=event_name,
                start_date=start_date,
                end_date=end_date,
                event_year=_as_int(
                    cell("event_year"), column="event_year", row_number=offset, allow_none=True
                ),
                event_month_key=_text(cell("event_month_key")),
                event_month_label=_text(cell("event_month_label")),
                event_quarter=_text(cell("event_quarter")),
                event_status=_text(cell("event_status")),
                tag_key=_text(cell("tag_key")),
                tag_label=_text(cell("tag_label")),
                event_type_key=_text(cell("event_type_key")),
                event_type_label=_text(cell("event_type_label")),
                delivery_mode=_text(cell("delivery_mode")),
                include_status=_text(cell("include_status")),
                member_price_eur=_as_money(
                    cell("member_price_eur"), column="member_price_eur", row_number=offset
                ),
                nonmember_price_eur=_as_money(
                    cell("nonmember_price_eur"), column="nonmember_price_eur", row_number=offset
                ),
                price_status=_text(cell("price_status")),
                added_date=_as_date(cell("added_date"), column="added_date", row_number=offset),
                planning_lead_days=_as_int(
                    cell("planning_lead_days"),
                    column="planning_lead_days",
                    row_number=offset,
                    allow_none=True,
                ),
                public_url=_text(cell("public_url"), limit=MAX_URL_LENGTH),
                public_link_status=_text(cell("public_link_status")),
                source_year=_as_int(cell("source_year"), column="source_year", row_number=offset),
                source_sheet=_text(cell("source_sheet")),
                source_row=_as_int(cell("source_row"), column="source_row", row_number=offset),
                source_occurrence_count=_as_int(
                    cell("source_occurrence_count"),
                    column="source_occurrence_count",
                    row_number=offset,
                ),
                date_parse_status=_text(cell("date_parse_status")),
                review_required=_as_bool(
                    cell("review_required"), column="review_required", row_number=offset
                ),
                warning_codes=_split_warning_codes(cell("warning_codes")),
            )
        )

    return tuple(rows)


def _require_agreement(control: WorkbookControl, rows: tuple[WorkbookRow, ...]) -> None:
    """The workbook's own summary must match the rows it actually carries.

    A file whose CONTROL sheet disagrees with its authoritative table is
    rejected rather than reconciled: one of the two is wrong, this importer
    cannot tell which, and guessing would publish a number nobody verified.
    """
    if control.canonical_event_count != len(rows):
        raise WorkbookContractError(
            f"DASH_CONTROL lubab {control.canonical_event_count} sündmust, "
            f"kuid tabelis on {len(rows)}."
        )

    linked = sum(1 for row in rows if row.public_url)
    if control.linked_public_url_count != linked:
        raise WorkbookContractError(
            f"DASH_CONTROL lubab {control.linked_public_url_count} avalikku linki, "
            f"kuid tabelis on {linked}."
        )


def parse_workbook(path: Path | str) -> ParsedWorkbook:
    """Read and validate the canonical event-programme workbook."""
    workbook = load_workbook(filename=str(path), data_only=True, read_only=False)
    try:
        _require_sheets(workbook)
        _require_tables(workbook)

        control = _require_control(_read_control(workbook[CONTROL_SHEET]))

        events_sheet = workbook[EVENTS_SHEET]
        header_row = _header_row_number(events_sheet, EVENTS_TABLE_NAME)
        _require_header(events_sheet, header_row)
        rows = _parse_rows(events_sheet, header_row)
    finally:
        workbook.close()

    _require_agreement(control, rows)
    return ParsedWorkbook(control=control, rows=rows)
