"""Import of aggregate composition facts from the member roster.

This is the module that touches personal data, so it is the module that has to
be right about not keeping any.

The roster is a spreadsheet with one row per member organisation, carrying the
company name, registry code, street address, director's name, two contact
addresses and a free-text comment. The dashboard needs none of that. It needs to
know how many members are in each size class, county, sector, tenure band and
joining year.

## How the boundary is enforced

Three separate things have to fail before an identity could be stored, and each
one is enough on its own:

1. **the reader never builds a record.** A row is streamed out of the workbook,
   six scalars are handed to `composition.build_member_row`, and the row is
   dropped on the next iteration of the loop. There is no list of members, no
   DataFrame and no intermediate file;
2. **the classifier returns buckets.** `MemberRow` has no name, code, address or
   comment field to carry one;
3. **the models cannot hold one.** `MembershipCompositionValue` stores a
   vocabulary key, a vocabulary label and an integer. There is no column an
   identity would fit in.

The same rule governs everything this module *says*. Diagnostics, the audit
summary, `--json` output and every exception message carry counts, column names
and vocabulary terms only. A parse failure reports the row number and the column
name, never the value that failed — a malformed registry code is still a
registry code.

## The snapshot date is an argument, not a filename

The workbook states no date of its own. Reading one out of the file name would
make a rename a data edit, and this repository has already been bitten by a file
whose name disagreed with its contents. So the date is supplied explicitly and
every tenure in the import is measured against it.

## Idempotency and revisions

The import key is the importer name, the schema version and the workbook's
SHA-256, so re-running the identical file reports `unchanged` and writes
nothing. A *different* file — a corrected export for the same date — is a
revision and needs `--supersede-previous`, which retires the existing snapshot
without deleting it. Neither its counts nor its checksum are touched, so what
the dashboard showed last month can still be traced to the file that produced
it.
"""

from __future__ import annotations

import hashlib
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from django.db import transaction

from apps.audit.services import record_event
from apps.sources.models import ImportRun, ImportStatus, SourceArtifact
from apps.sources.services import (
    build_import_run,
    calculate_import_key,
    complete_import_run,
    fail_publication,
    register_external_reference,
    start_import_run,
)

from .audit_actions import MembershipAudit
from .bootstrap import ensure_membership_composition_source
from .composition import (
    MEMBERSHIP_COMPOSITION_MAPPING_VERSION,
    MEMBERSHIP_SECTOR_MAPPING_VERSION,
    CompositionTally,
    Dimension,
    Population,
    build_member_row,
    category_label,
)
from .models import MembershipCompositionSnapshot, MembershipCompositionValue

IMPORTER_NAME = "membership_composition_xlsx"
SCHEMA_VERSION = "1.0"

# A fixed, non-secret provenance label. It names what the content was, carries
# no credential and no path, and satisfies the artifact model's rule that an
# external reference contains neither `@` nor `?`.
ARTIFACT_REFERENCE_PREFIX = "roster:membership-composition"

# The roster is about 750 kB. Fifty is room for the membership to grow several
# times over and still refuses a file that is not a roster at all, before
# openpyxl is asked to open it.
MAX_SOURCE_BYTES = 50 * 1024 * 1024

# The columns this importer reads, by their heading in the roster.
#
# Everything else in the workbook — the company name, the address, the postal
# index, the town, the two e-mail columns, the website, the director, the
# registry code, the free-text comment and the NACE comment — is deliberately
# not listed, and a column that is not listed is never read.
COLUMN_STATUS = "Staatus"
COLUMN_LEGAL_FORM = "Vorm"
COLUMN_REGION = "Maakond"
COLUMN_EMPLOYEES = "Töötajate arv"
COLUMN_START = "Algus kp."
COLUMN_SECTOR = "Nace kood"

REQUIRED_COLUMNS: tuple[str, ...] = (
    COLUMN_STATUS,
    COLUMN_LEGAL_FORM,
    COLUMN_REGION,
    COLUMN_EMPLOYEES,
    COLUMN_START,
    COLUMN_SECTOR,
)

# `Töötaja vahemik` is deliberately not read, and this is the reason rather than
# an oversight: Excel has coerced two thirds of its values into dates, so
# `1-4` arrives as a timestamp. The integer `Töötajate arv` beside it is
# complete and unambiguous, and the bands in `composition.py` are derived from
# it. Reading a column that a spreadsheet has silently corrupted would put
# invented size classes on the page.
COLUMN_CORRUPTED_BY_EXCEL = "Töötaja vahemik"


class CompositionImportError(RuntimeError):
    """A refusal that names a column or a count, never a cell value."""


@dataclass(frozen=True)
class CompositionImportResult:
    import_run: ImportRun
    dry_run: bool
    unchanged: bool
    source_sha256: str
    snapshot_date: date | None
    rows_read: int = 0
    values_written: int = 0
    superseded: int = 0
    diagnostics: tuple[dict, ...] = field(default_factory=tuple)

    def as_json(self) -> dict:
        """Aggregate counts and identifiers only. Never source content."""
        return {
            "importer": IMPORTER_NAME,
            "schema_version": SCHEMA_VERSION,
            "import_run_id": self.import_run.pk if self.import_run else None,
            "dry_run": self.dry_run,
            "unchanged": self.unchanged,
            "source_sha256": self.source_sha256,
            "snapshot_date": self.snapshot_date.isoformat() if self.snapshot_date else None,
            "rows_read": self.rows_read,
            "values_written": self.values_written,
            "superseded_snapshots": self.superseded,
            "diagnostics": list(self.diagnostics),
        }


def _checksum(path: Path) -> tuple[str, int]:
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
            size += len(block)
    return digest.hexdigest(), size


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def read_roster(path: Path | str, *, snapshot_date: date) -> CompositionTally:
    """Stream the roster and return counts. No row survives the loop.

    Raises `CompositionImportError` when the file is not a roster this importer
    recognises — a missing required column is refused rather than silently
    producing a page full of `Teadmata`.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - dependency is declared
        raise CompositionImportError("openpyxl ei ole saadaval.") from error

    path = Path(path)
    if not path.is_file():
        raise CompositionImportError("Lähtefaili ei leitud.")
    if path.stat().st_size > MAX_SOURCE_BYTES:
        raise CompositionImportError("Lähtefail on lubatust suurem.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as error:
            raise CompositionImportError("Lähtefail on tühi.") from error

        columns = {
            str(name).strip(): index for index, name in enumerate(header) if name is not None
        }
        missing = [name for name in REQUIRED_COLUMNS if name not in columns]
        if missing:
            # Column names are structure, not content, so naming them is safe
            # and is the only way the failure is actionable.
            raise CompositionImportError("Lähtefailis puuduvad veerud: " + ", ".join(missing))

        tally = CompositionTally(snapshot_date=snapshot_date)
        future_starts = 0
        unreadable_starts = 0

        for row in rows:
            if all(
                cell is None or (isinstance(cell, str) and not cell.strip())
                for cell in row[: len(header)]
            ):
                continue

            def cell(name: str):
                index = columns[name]
                return row[index] if index < len(row) else None

            start = _as_date(cell(COLUMN_START))
            if cell(COLUMN_START) is not None and start is None:
                unreadable_starts += 1
            if start is not None and start > snapshot_date:
                future_starts += 1

            tally.add(
                build_member_row(
                    status=cell(COLUMN_STATUS),
                    legal_form=cell(COLUMN_LEGAL_FORM),
                    employees=cell(COLUMN_EMPLOYEES),
                    region=cell(COLUMN_REGION),
                    sector_code=cell(COLUMN_SECTOR),
                    membership_start=start,
                    snapshot_date=snapshot_date,
                )
            )
            # `row` is rebound on the next iteration and nothing holds a
            # reference to it. This is the point at which the member is gone.

        if not tally.rows_read:
            raise CompositionImportError("Lähtefailis ei ole ühtegi andmerida.")

        tally.unmapped["_future_start_dates"] = future_starts
        tally.unmapped["_unreadable_start_dates"] = unreadable_starts
        return tally
    finally:
        workbook.close()


def validate(tally: CompositionTally) -> list[dict]:
    """Structural checks, returning diagnostics that name no member.

    A violation of an invariant this importer controls raises; a fact about the
    source's own quality is reported as a diagnostic and does not block the
    import. The difference matters: a roster with a few unclassified sectors is
    a roster, and refusing it would leave the dashboard with nothing rather than
    with a measured gap.
    """
    diagnostics: list[dict] = []
    total = tally.total(Population.ALL_CURRENT)

    if total != tally.rows_read:
        raise CompositionImportError(
            f"Ridade arv ei klapi: loetud {tally.rows_read}, liigitatud {total}."
        )

    recent_total = tally.total(Population.RECENT_JOINERS)
    if recent_total > total:
        raise CompositionImportError(
            f"Hiljuti liitunuid ({recent_total}) on rohkem kui liikmeid kokku ({total})."
        )

    for dimension in (
        Dimension.STATUS,
        Dimension.LEGAL_FORM,
        Dimension.EMPLOYEE_SIZE,
        Dimension.REGION,
        Dimension.SECTOR,
        Dimension.TENURE_BAND,
        Dimension.JOIN_COHORT,
    ):
        counted = sum(tally.category_counts(Population.ALL_CURRENT, dimension).values())
        if counted != total:
            raise CompositionImportError(
                f"Mõõde {dimension} ei kata kõiki ridu: {counted} / {total}."
            )
        coverage = tally.coverage_pct(dimension)
        if coverage is not None and coverage < 100:
            diagnostics.append(
                {
                    "code": "unclassified_values",
                    "dimension": dimension,
                    "unclassified": tally.unmapped.get(dimension, 0),
                    "coverage_pct": str(coverage),
                }
            )

    for code, key in (
        ("membership_start_after_snapshot", "_future_start_dates"),
        ("unreadable_membership_start", "_unreadable_start_dates"),
    ):
        count = tally.unmapped.get(key, 0)
        if count:
            diagnostics.append({"code": code, "rows": count})

    return diagnostics


def _ensure_artifact(source, *, sha256: str, size: int, actor, correlation_id) -> SourceArtifact:
    """Register the reading's identity without keeping the file.

    The roster holds personal data, so storing it would put a member list on a
    served path and inside every backup. An artifact is importable when it has a
    trusted checksum, not when it still has a file — the same rule the historical
    package import follows, and here it is a privacy control as well as a
    storage one.
    """
    return register_external_reference(
        source=source,
        external_reference=f"{ARTIFACT_REFERENCE_PREFIX}:{sha256}",
        sha256=sha256,
        size_bytes=size,
        mime_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        actor=actor,
        correlation_id=correlation_id,
    )


def _existing_successful_run(import_key: str) -> ImportRun | None:
    return (
        ImportRun.objects.filter(
            import_key=import_key, status=ImportStatus.SUCCEEDED, dry_run=False
        )
        .order_by("-id")
        .first()
    )


def _retire_current(source, *, supersede_previous: bool) -> list[MembershipCompositionSnapshot]:
    """Stand down the snapshot in force, or refuse to import over it.

    A second roster is not an accident to be absorbed silently: it either
    corrects the one on the page or it is a mistake, and only the operator knows
    which. `--supersede-previous` is how they say so.
    """
    current = list(
        MembershipCompositionSnapshot.objects.select_for_update().filter(
            source=source, is_current=True
        )
    )
    if not current:
        return []
    if not supersede_previous:
        raise CompositionImportError(
            "Koosseisu hetkeseis on juba olemas. Uue importimiseks kasuta "
            "--supersede-previous, mis märgib varasema asendatuks. Midagi ei kustutata."
        )
    return current


def _write_values(snapshot: MembershipCompositionSnapshot, tally: CompositionTally) -> int:
    rows = []
    for population in (Population.ALL_CURRENT, Population.RECENT_JOINERS):
        for dimension in (
            Dimension.STATUS,
            Dimension.LEGAL_FORM,
            Dimension.EMPLOYEE_SIZE,
            Dimension.REGION,
            Dimension.SECTOR,
            Dimension.TENURE_BAND,
            Dimension.JOIN_COHORT,
        ):
            for key, count in tally.category_counts(population, dimension).items():
                rows.append(
                    MembershipCompositionValue(
                        snapshot=snapshot,
                        population=population,
                        dimension=dimension,
                        category_key=key,
                        category_label=category_label(dimension, key),
                        member_count=count,
                    )
                )
    MembershipCompositionValue.objects.bulk_create(rows, batch_size=500)
    return len(rows)


def import_composition_snapshot(
    path: Path | str,
    *,
    snapshot_date: date,
    dry_run: bool = True,
    supersede_previous: bool = False,
    actor=None,
    correlation_id: uuid.UUID | None = None,
) -> CompositionImportResult:
    """Read the roster, aggregate it, and store the aggregates.

    A dry run reads and validates the whole workbook, records the attempt and
    writes no domain row — which is also the way to check what a new export
    would produce before it replaces what is on the page.
    """
    path = Path(path)
    if not path.is_file():
        raise CompositionImportError("Lähtefaili ei leitud.")

    sha256, size = _checksum(path)
    source = ensure_membership_composition_source(actor=actor, correlation_id=correlation_id)
    import_key = calculate_import_key(IMPORTER_NAME, SCHEMA_VERSION, sha256)

    already = _existing_successful_run(import_key)
    if already is not None and not dry_run:
        record_event(
            action=MembershipAudit.COMPOSITION_UNCHANGED,
            obj=already,
            actor=actor,
            correlation_id=already.correlation_id,
            change_summary={
                "source": source.slug,
                "source_sha256": sha256,
                "import_key": import_key,
            },
        )
        return CompositionImportResult(
            import_run=already,
            dry_run=False,
            unchanged=True,
            source_sha256=sha256,
            snapshot_date=snapshot_date,
        )

    tally = read_roster(path, snapshot_date=snapshot_date)
    diagnostics = validate(tally)

    artifact = _ensure_artifact(
        source, sha256=sha256, size=size, actor=actor, correlation_id=correlation_id
    )
    run = build_import_run(
        artifact=artifact,
        importer_name=IMPORTER_NAME,
        schema_version=SCHEMA_VERSION,
        dry_run=dry_run,
        initiated_by=actor,
        actor=actor,
        correlation_id=correlation_id,
    )
    start_import_run(run)

    try:
        if dry_run:
            complete_import_run(
                run,
                rows_added=0,
                rows_skipped=tally.rows_read,
                warnings=diagnostics,
                actor=actor,
            )
            return CompositionImportResult(
                import_run=run,
                dry_run=True,
                unchanged=False,
                source_sha256=sha256,
                snapshot_date=snapshot_date,
                rows_read=tally.rows_read,
                diagnostics=tuple(diagnostics),
            )

        with transaction.atomic():
            retiring = _retire_current(source, supersede_previous=supersede_previous)
            snapshot = MembershipCompositionSnapshot.objects.create(
                source=source,
                import_run=run,
                snapshot_date=snapshot_date,
                source_sha256=sha256,
                source_row_count=tally.rows_read,
                mapping_version=MEMBERSHIP_COMPOSITION_MAPPING_VERSION,
                sector_mapping_version=MEMBERSHIP_SECTOR_MAPPING_VERSION,
                median_tenure_days=tally.median_tenure_days,
                coverage_pct={
                    dimension: str(tally.coverage_pct(dimension) or Decimal(0))
                    for dimension in (
                        Dimension.STATUS,
                        Dimension.LEGAL_FORM,
                        Dimension.EMPLOYEE_SIZE,
                        Dimension.REGION,
                        Dimension.SECTOR,
                        Dimension.TENURE_BAND,
                        Dimension.JOIN_COHORT,
                    )
                },
                is_current=True,
            )
            for previous in retiring:
                previous.is_current = False
                previous.superseded_by = snapshot
                previous.save(update_fields=["is_current", "superseded_by"])

            written = _write_values(snapshot, tally)

            complete_import_run(
                run,
                rows_added=written + 1,
                rows_skipped=0,
                warnings=diagnostics,
                actor=actor,
            )
            record_event(
                action=MembershipAudit.COMPOSITION_IMPORTED,
                obj=snapshot,
                actor=actor,
                correlation_id=run.correlation_id,
                change_summary={
                    "source": source.slug,
                    "source_sha256": sha256,
                    "snapshot_date": snapshot_date.isoformat(),
                    "rows_read": tally.rows_read,
                    "values_written": written,
                    "superseded_snapshots": len(retiring),
                    "mapping_version": MEMBERSHIP_COMPOSITION_MAPPING_VERSION,
                },
            )

        return CompositionImportResult(
            import_run=run,
            dry_run=False,
            unchanged=False,
            source_sha256=sha256,
            snapshot_date=snapshot_date,
            rows_read=tally.rows_read,
            values_written=written,
            superseded=len(retiring),
            diagnostics=tuple(diagnostics),
        )
    except Exception as error:
        # The message is the exception's own text, which every raise in this
        # module keeps free of cell values.
        fail_publication(run, errors=[{"code": "composition_import_failed"}], actor=actor)
        if isinstance(error, CompositionImportError):
            raise
        raise CompositionImportError("Koosseisu import ebaõnnestus.") from error
