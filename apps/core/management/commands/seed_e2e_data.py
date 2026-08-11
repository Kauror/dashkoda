"""Publish deterministic synthetic content for the browser acceptance suite.

CI's database is empty, so the browser suite has always exercised empty states
only. That is why a real 152-pixel horizontal overflow reached production while
every viewport assertion passed: nothing in CI was ever long enough to truncate.
This command fills the database with content shaped to expose exactly that class
of defect — very long Estonian titles, linked headings carrying a visually
hidden suffix, wide amounts, explicit zeros beside genuinely missing values, and
enough rows to scroll.

**Every value here is invented.** No Chamber member total, fee figure, event,
legal topic, article, organisation or URL appears. The names are obviously
synthetic so that a screenshot can never be mistaken for real data.

It publishes through the domain services rather than writing rows directly, so
the seeded state is one the application could actually have reached: the same
collectors, importers, import registry, atomic publication and audit trail. No
immutability guard is weakened to make seeding easier, and nothing here performs
a network request or touches a real source.

Re-running is safe. Every publisher is idempotent over its own content identity
— the feed syncs by checksum, the manual publishers by content hash — so a
second run on the same day publishes nothing new and reports `unchanged`.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import os
import re
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

# Settings modules under which seeding is permitted. Production is refused by
# construction rather than by a flag someone can pass anyway.
ALLOWED_SETTINGS_MODULES = frozenset(
    {
        "config.settings.local",
        "config.settings.test",
    }
)

# One long Estonian sentence, reused where a title has to be long enough to
# truncate. Long enough to overflow a narrow card, and unmistakably synthetic.
LONG_TITLE = (
    "Sünteetiline väga pikk pealkiri, mis on kirjutatud ainult selleks, "
    "et kontrollida kärpimist, murdmist ja horisontaalset kerimist kõige "
    "kitsamas vaates, ning see ei kirjelda ühtegi tegelikku Koja tegevust"
)
LONG_TOPIC = (
    "Sünteetiline õigusloome teema, mille pealkiri on tahtlikult äärmiselt pikk, "
    "et kontrollida tabeliveeru kärpimist ja seda, kas pikk seotud pealkiri koos "
    "peidetud lisamärkusega ajab lehe horisontaalselt kerima; ükski sõna siin ei "
    "puuduta tegelikku õigusloomet ega Koja seisukohti"
)
LONG_LOCATION = "Sünteetiline konverentsikeskus, sünteetiline suur saal, sünteetiline aadress 123"
LONG_CATEGORY = "Sünteetiline pikk kategooria nimetus"


def _require_non_production() -> str:
    """Refuse to seed anything but an explicit development or test database."""
    module = os.environ.get("DJANGO_SETTINGS_MODULE", "")
    if module not in ALLOWED_SETTINGS_MODULES:
        raise CommandError(
            "seed_e2e_data refuses to run under "
            f"{module or '(unset DJANGO_SETTINGS_MODULE)'}. "
            "It is permitted only under " + ", ".join(sorted(ALLOWED_SETTINGS_MODULES)) + "."
        )
    return module


# --------------------------------------------------------------------------
# Legal work — a real workbook through the real parser
# --------------------------------------------------------------------------


def _legal_work_rows(today: dt.date) -> list[list]:
    """DATA rows in canonical column order, deliberately varied.

    Covers a long topic, a topic with no deadline, all three sent statuses, a
    warning code, and enough open rows that the page's list has to scroll.
    """
    from apps.legal_work.workbook import DATA_COLUMNS  # noqa: F401  (column order documented)

    rows: list[list] = []

    def row(
        *,
        index: int,
        topic: str,
        received_offset: int,
        deadline_offset: int | None,
        sent_offset: int | None = None,
        sent_status: str = "pending",
        is_open: bool = True,
        warning_codes: str | None = None,
        stage: str = "sünteetiline menetlusetapp",
    ) -> list:
        return [
            f"SEED-{index:04d}",
            2099,
            index,
            topic,
            "sünteetiline õigusakti liik",
            today - dt.timedelta(days=received_offset),
            None if deadline_offset is None else today + dt.timedelta(days=deadline_offset),
            None if sent_offset is None else today - dt.timedelta(days=sent_offset),
            sent_status,
            "Sünteetiline saaja ministeerium",
            stage,
            stage,
            "sünteetiline järgmine samm",
            is_open,
            warning_codes,
            index + 1,
            dt.datetime.combine(today - dt.timedelta(days=1), dt.time(6, 30)),
        ]

    # The overflow candidate: an extremely long topic in a linked table cell.
    rows.append(row(index=1, topic=LONG_TOPIC, received_offset=40, deadline_offset=2))
    # Deadlines across the urgency thresholds the selectors use.
    rows.append(
        row(
            index=2,
            topic="Sünteetiline kiireloomuline teema",
            received_offset=30,
            deadline_offset=1,
        )
    )
    rows.append(
        row(index=3, topic="Sünteetiline lähituleviku teema", received_offset=25, deadline_offset=8)
    )
    rows.append(
        row(index=4, topic="Sünteetiline rahulik teema", received_offset=20, deadline_offset=18)
    )
    # Open, but with no deadline at all: must sort last, never first.
    rows.append(
        row(index=5, topic="Sünteetiline tähtajata teema", received_offset=15, deadline_offset=None)
    )
    # A warning code, and an empty stage, which is what produces `missing_stage`.
    rows.append(
        row(
            index=6,
            topic="Sünteetiline hoiatusega teema",
            received_offset=12,
            deadline_offset=5,
            warning_codes="missing_stage",
            stage="",
        )
    )
    # Sent opinions, so "viimati välja läinud" has content.
    for offset, index in enumerate(range(7, 13)):
        rows.append(
            row(
                index=index,
                topic=f"Sünteetiline saadetud arvamus {index}",
                received_offset=60 + offset,
                deadline_offset=None,
                sent_offset=5 + offset,
                sent_status="sent",
                is_open=False,
                stage="jõustunud",
            )
        )
    # Explicitly not sent: distinct from pending, and carries no date.
    rows.append(
        row(
            index=13,
            topic="Sünteetiline saatmata jäetud teema",
            received_offset=70,
            deadline_offset=None,
            sent_status="not_sent",
            is_open=False,
            stage="lõpetatud",
        )
    )
    # Enough remaining open rows that the bounded list actually scrolls.
    for index in range(14, 25):
        rows.append(
            row(
                index=index,
                topic=f"Sünteetiline töös olev teema {index}",
                received_offset=index,
                deadline_offset=index,
            )
        )
    return rows


# An XLSX carries the current time in two independent places, and both have to
# be frozen or the seed publishes a fresh snapshot on every run: the ZIP member
# headers, and the `dcterms:created` / `dcterms:modified` fields openpyxl writes
# into `docProps/core.xml`. Freezing only the first looks like it works, because
# two builds inside the same second still hash identically — it fails as soon as
# they straddle a second boundary. Both values are far in the future and
# obviously synthetic.
FIXED_ZIP_TIMESTAMP = (2099, 1, 1, 0, 0, 0)
FIXED_DOCUMENT_TIMESTAMP = "2099-01-01T00:00:00Z"


CORE_PROPERTIES_MEMBER = "docProps/core.xml"
_DOCUMENT_TIMESTAMP_PATTERN = re.compile(
    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)"
)


def _freeze_package_timestamps(path: Path) -> None:
    """Rewrite the package so identical content produces identical bytes.

    Both timestamps are handled here rather than on the workbook object,
    because openpyxl re-stamps ``dcterms:modified`` with the current time while
    saving — assigning it beforehand looks like it works and silently does not.
    Doing it in one rewrite pass keeps a single mechanism for a single job.
    """
    import zipfile

    with zipfile.ZipFile(path) as source:
        members = [(info, source.read(info.filename)) for info in source.infolist()]

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in members:
            if info.filename == CORE_PROPERTIES_MEMBER:
                payload = _DOCUMENT_TIMESTAMP_PATTERN.sub(
                    rb"\g<1>" + FIXED_DOCUMENT_TIMESTAMP.encode("ascii") + rb"\g<2>",
                    payload,
                )
            frozen = zipfile.ZipInfo(info.filename, date_time=FIXED_ZIP_TIMESTAMP)
            frozen.compress_type = zipfile.ZIP_DEFLATED
            frozen.external_attr = info.external_attr
            target.writestr(frozen, payload)


def _write_legal_work_workbook(path: Path, today: dt.date) -> Path:
    """Write a workbook that satisfies the canonical contract exactly."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    from apps.legal_work.workbook import (
        CONTROL_SHEET,
        DATA_COLUMNS,
        DATA_SHEET,
        DATA_TABLE_NAME,
        OVERVIEW_SHEET,
        WARNINGS_SHEET,
    )

    rows = _legal_work_rows(today)
    is_open_index = DATA_COLUMNS.index("is_open")
    sent_status_index = DATA_COLUMNS.index("sent_status")
    warnings_index = DATA_COLUMNS.index("warning_codes")
    reporting_date = today - dt.timedelta(days=1)
    generated_at = dt.datetime.combine(reporting_date, dt.time(6, 30))

    control = {
        "dataset_key": "oigusloome",
        "schema_version": "1.1",
        "source_file_name": "sünteetiline-lähtefail.xlsx",
        "source_sheet": "2099",
        "source_modified_at": "",
        "source_sha256": "0" * 64,
        "generated_at": generated_at.isoformat(),
        "reporting_date": reporting_date,
        "total_record_count": len(rows),
        "open_record_count": sum(1 for row in rows if row[is_open_index] is True),
        "sent_record_count": sum(1 for row in rows if row[sent_status_index] == "sent"),
        "not_sent_record_count": sum(1 for row in rows if row[sent_status_index] == "not_sent"),
        "warning_record_count": sum(1 for row in rows if row[warnings_index]),
        "refresh_status": "completed_with_warnings",
        "generator_version": "seed-e2e",
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in (CONTROL_SHEET, OVERVIEW_SHEET, DATA_SHEET, WARNINGS_SHEET):
        workbook.create_sheet(name)

    control_sheet = workbook[CONTROL_SHEET]
    control_sheet.append(["DASHKODA ÕIGUSLOOME ANDMEVOOG (SÜNTEETILINE)", None])
    for key, value in control.items():
        control_sheet.append([key, value])

    workbook[OVERVIEW_SHEET].append(["Sünteetiline ülevaade"])

    data = workbook[DATA_SHEET]
    data.append(list(DATA_COLUMNS))
    for row in rows:
        data.append(row)
    last_column = chr(ord("A") + len(DATA_COLUMNS) - 1)
    data.add_table(Table(displayName=DATA_TABLE_NAME, ref=f"A1:{last_column}{len(rows) + 1}"))

    warnings_sheet = workbook[WARNINGS_SHEET]
    warnings_sheet.append(["record_id", "source_row", "field", "warning_code"])
    for row in rows:
        if not row[warnings_index]:
            continue
        for code in str(row[warnings_index]).split(";"):
            warnings_sheet.append([row[0], row[-2], "stage", code.strip()])
    warnings_sheet.add_table(
        Table(displayName="tbl_oigusloome_warnings", ref=f"A1:D{warnings_sheet.max_row}")
    )

    workbook.save(path)
    _freeze_package_timestamps(path)
    return path


def _seed_legal_work(today: dt.date) -> str:
    from apps.legal_work.public_download import XLSX_MIME_TYPE, PublicDownload
    from apps.legal_work.public_sync import synchronize_public_workbook
    from apps.legal_work.sync import SyncLocked, advisory_lock

    def downloader(destination: Path) -> PublicDownload:
        _write_legal_work_workbook(destination, today)
        payload = destination.read_bytes()
        return PublicDownload(
            path=destination,
            size_bytes=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            content_type=XLSX_MIME_TYPE,
            final_host="synthetic-seed.invalid",
        )

    try:
        with advisory_lock():
            outcome = synchronize_public_workbook(downloader=downloader)
    except SyncLocked as error:
        raise CommandError(f"Legal-work seed could not take the lock: {error}") from error
    return f"õigusloome: {outcome.result} ({outcome.rows_imported} kirjet)"


# --------------------------------------------------------------------------
# Event programme — a real workbook through the real parser and importer
# --------------------------------------------------------------------------

# A synthetic public page on the allowed host. `/et/sundmused/` is the real
# prefix, and `sunteetiline-*` is unmistakably not a production path, so a
# screenshot can never be read as a link to a real Chamber event.
SYNTHETIC_EVENT_URL = "https://www.koda.ee/et/sundmused/sunteetiline-programmi-sundmus"

LONG_EVENT_NAME = (
    "Sünteetiline väga pikk sündmuse nimi, mis on kirjutatud ainult selleks, et "
    "kontrollida tabeliveeru kärpimist ja seda, kas pikk seotud pealkiri koos "
    "peidetud lisamärkusega ajab lehe horisontaalselt kerima; see ei kirjelda "
    "ühtegi tegelikku Koja sündmust"
)

# Estonian month names, as the Chamber's generator writes `event_month_label`.
EVENT_MONTH_LABELS = (
    "jaanuar",
    "veebruar",
    "märts",
    "aprill",
    "mai",
    "juuni",
    "juuli",
    "august",
    "september",
    "oktoober",
    "november",
    "detsember",
)

# Deliberately several of each, so tag and type filtering has something to
# separate and the option lists are not single-valued.
EVENT_TAGS = (
    ("seminar", "Sünteetiline seminar"),
    ("konverents", "Sünteetiline konverents"),
    ("koolitus", "Sünteetiline koolitus"),
)
EVENT_TYPES = (
    ("training", "Sünteetiline koolitusvorm"),
    ("conference", "Sünteetiline konverentsivorm"),
)
DELIVERY_MODES = ("onsite", "online", "hybrid")


def _programme_row(
    *,
    index: int,
    name: str,
    start: dt.date | None,
    end: dt.date | None = None,
    status: str,
    tag: tuple[str, str],
    event_type: tuple[str, str],
    delivery_mode: str,
    public_url: str | None = None,
    review_required: bool = False,
) -> dict:
    """One synthetic `DASH_EVENTS` row, keyed by the contract's column names.

    The four derived calendar fields come from `start` exactly as the generator
    derives them, so an undated row carries no year, month or quarter rather than
    an invented one.
    """
    tag_key, tag_label = tag
    type_key, type_label = event_type
    dated = start is not None
    return {
        "event_id": f"SEED-EVENT-{index:04d}",
        "service_code": f"S{index:04d}",
        "event_name": name,
        "event_name_raw": name,
        "date_text_raw": start.isoformat() if dated else "sünteetiline kuupäevatekst",
        "start_date": dt.datetime.combine(start, dt.time()) if dated else None,
        "end_date": dt.datetime.combine(end or start, dt.time()) if dated else None,
        "event_year": start.year if dated else None,
        "event_month": start.month if dated else None,
        "event_month_key": f"{start:%Y-%m}" if dated else None,
        "event_month_label": EVENT_MONTH_LABELS[start.month - 1] if dated else None,
        "event_quarter": f"Q{(start.month - 1) // 3 + 1}" if dated else None,
        "event_status": status,
        "short_name_raw": "SÜN",
        "short_name_normalized": "syn",
        "tag_key": tag_key,
        "tag_label": tag_label,
        "event_type_key": type_key,
        "event_type_label": type_label,
        "delivery_mode": delivery_mode,
        "include_status": "REVIEW" if review_required else "YES",
        "group_raw": None,
        "group_secondary_raw": None,
        # Pricing exists in the export and must never reach a model field.
        "member_price_raw": "100",
        "member_price_eur": 100,
        "nonmember_price_raw": "200",
        "nonmember_price_eur": 200,
        "later_member_price_raw": None,
        "later_member_price_eur": None,
        "later_nonmember_price_raw": None,
        "later_nonmember_price_eur": None,
        "price_status": "parsed",
        "discount_code": None,
        "discount_raw": None,
        "added_date": None,
        "planning_lead_days": None,
        "public_url": public_url,
        "public_link_status": "linked_embedded_latest" if public_url else "not_linked",
        "source_year": start.year if dated else 2099,
        # Deliberately not the event's own year: the page must never present the
        # annual sheet a row sat on as the event's date.
        "source_sheet": "KOOD 2099",
        "source_row": index + 1,
        "source_occurrence_count": 1,
        "date_parse_status": "parsed_range"
        if dated and end and end != start
        else ("parsed_single" if dated else "unparsed"),
        "review_required": review_required,
        "warning_codes": "date_unparsed" if not dated else None,
        "export_refreshed_at": None,
    }


def _programme_rows(today: dt.date) -> list[dict]:
    """A programme spanning several years, with every awkward shape in it.

    Covers past, ongoing and upcoming events, a multi-day range, a quarter
    boundary, an undated event, three tags, two types, all three delivery modes, a
    linked and an unlinked title, a review-required record, a very long linked
    name and enough rows to need more than one page.
    """
    rows: list[dict] = []
    index = 0

    def add(**kwargs) -> None:
        nonlocal index
        index += 1
        rows.append(_programme_row(index=index, **kwargs))

    def tag(position: int) -> tuple[str, str]:
        return EVENT_TAGS[position % len(EVENT_TAGS)]

    def event_type(position: int) -> tuple[str, str]:
        return EVENT_TYPES[position % len(EVENT_TYPES)]

    # The overflow candidate: a very long linked name in a table cell.
    add(
        name=LONG_EVENT_NAME,
        start=today + dt.timedelta(days=9),
        status="upcoming",
        tag=tag(0),
        event_type=event_type(0),
        delivery_mode="onsite",
        public_url=SYNTHETIC_EVENT_URL,
    )
    # Under way right now: started before today and ends after it.
    add(
        name="Sünteetiline käimasolev sündmus",
        start=today - dt.timedelta(days=1),
        end=today + dt.timedelta(days=1),
        status="ongoing",
        tag=tag(1),
        event_type=event_type(1),
        delivery_mode="hybrid",
    )
    # Inside the backward 30-day window, and unlinked.
    add(
        name="Sünteetiline hiljuti toimunud sündmus",
        start=today - dt.timedelta(days=10),
        status="past",
        tag=tag(2),
        event_type=event_type(0),
        delivery_mode="online",
    )
    # Inside the forward 30-day window, as a multi-day range.
    add(
        name="Sünteetiline mitmepäevane sündmus",
        start=today + dt.timedelta(days=3),
        end=today + dt.timedelta(days=5),
        status="upcoming",
        tag=tag(0),
        event_type=event_type(1),
        delivery_mode="onsite",
    )
    # The quarter boundary: 31 March and 1 April of the same year.
    boundary_year = today.year - 1
    add(
        name="Sünteetiline kvartali lõpu sündmus",
        start=dt.date(boundary_year, 3, 31),
        status="past",
        tag=tag(1),
        event_type=event_type(0),
        delivery_mode="online",
    )
    add(
        name="Sünteetiline kvartali alguse sündmus",
        start=dt.date(boundary_year, 4, 1),
        status="past",
        tag=tag(2),
        event_type=event_type(1),
        delivery_mode="hybrid",
    )
    # A real event whose operational row held date text nobody could parse. It
    # must stay a record, must not acquire a date, and must be reachable.
    add(
        name="Sünteetiline kuupäevata sündmus",
        start=None,
        status="date_unknown",
        tag=tag(0),
        event_type=event_type(0),
        delivery_mode="onsite",
        review_required=True,
    )
    # Enough history that year filtering has several years to choose between and
    # that the whole programme runs past one page of 50 rows.
    for offset in range(1, 4):
        year = today.year - offset
        for month in (2, 3, 5, 9, 11, 12):
            for day in (7, 14, 21):
                add(
                    name=f"Sünteetiline sündmus {year}-{month:02d}-{day:02d}",
                    start=dt.date(year, month, day),
                    status="past",
                    tag=tag(month + day),
                    event_type=event_type(month),
                    delivery_mode=DELIVERY_MODES[(month + day) % len(DELIVERY_MODES)],
                )
    return rows


def _write_event_programme_workbook(path: Path, today: dt.date) -> Path:
    """Write a workbook that satisfies the canonical event contract exactly.

    Built here rather than imported from the test suite: this command ships in the
    application image and may not depend on `tests/`. It writes the same six
    sheets, the same two-row banner and the same text CONTROL table the Chamber's
    Office Script produces, so the real parser is genuinely exercised.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    from apps.event_programme.workbook import (
        CONTROL_SHEET,
        CONTROL_TABLE_NAME,
        EVENTS_COLUMNS,
        EVENTS_SHEET,
        EVENTS_TABLE_NAME,
        OCCURRENCES_SHEET,
        REVIEW_SHEET,
        TAG_MAP_SHEET,
        URL_OVERRIDES_SHEET,
    )

    # The banner rows the generator writes above every DASH_* table.
    banner_rows = 2
    control_banner_rows = 4

    rows = _programme_rows(today)
    export_refreshed_at = dt.datetime.combine(today, dt.time(6, 30)).isoformat()
    for row in rows:
        row["export_refreshed_at"] = export_refreshed_at

    linked = sum(1 for row in rows if row["public_url"])
    control = {
        "dataset_key": "events",
        "schema_version": "1.0",
        "generator_name": "Sünteetiline seemnegeneraator",
        "generator_version": "seed-e2e",
        "source_workbook_name": "sünteetiline-lähtefail.xlsx",
        "refresh_status": "ready_with_warnings",
        "canonical_event_count": str(len(rows)),
        "qualifying_occurrence_count": str(len(rows)),
        "excluded_event_count": "0",
        "repeated_service_code_count": "0",
        "linked_public_url_count": str(linked),
        "distinct_short_name_count": "1",
        "blocking_error_count": "0",
        "warning_count": str(sum(1 for row in rows if row["warning_codes"])),
        "export_refreshed_at": export_refreshed_at,
        "last_successful_refresh_at": export_refreshed_at,
    }

    def write_table(sheet, *, table_name, columns, table_rows, banners) -> None:
        for offset in range(banners):
            sheet.cell(row=offset + 1, column=1, value=f"Sünteetiline bänner {offset + 1}")
        header_row = banners + 1
        for position, name in enumerate(columns, start=1):
            sheet.cell(row=header_row, column=position, value=name)
        for row_offset, values in enumerate(table_rows, start=header_row + 1):
            for position, value in enumerate(values, start=1):
                sheet.cell(row=row_offset, column=position, value=value)
        last_row = header_row + max(len(table_rows), 1)
        last_column = sheet.cell(row=header_row, column=len(columns)).column_letter
        sheet.add_table(Table(displayName=table_name, ref=f"A{header_row}:{last_column}{last_row}"))

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_table(
        workbook.create_sheet(CONTROL_SHEET),
        table_name=CONTROL_TABLE_NAME,
        columns=("key", "value"),
        table_rows=[(key, value) for key, value in control.items()],
        banners=control_banner_rows,
    )
    write_table(
        workbook.create_sheet(EVENTS_SHEET),
        table_name=EVENTS_TABLE_NAME,
        columns=EVENTS_COLUMNS,
        table_rows=[tuple(row.get(name) for name in EVENTS_COLUMNS) for row in rows],
        banners=banner_rows,
    )
    # Present with their tables, never read as a data source.
    for sheet_name, table_name, columns in (
        (OCCURRENCES_SHEET, "tbl_dash_event_occurrences", ("event_id", "occurrence_number")),
        (REVIEW_SHEET, "tbl_dash_review", ("severity", "issue_code")),
        (TAG_MAP_SHEET, "tbl_dash_tag_map", ("short_name_normalized", "tag_key")),
        (URL_OVERRIDES_SHEET, "tbl_dash_url_overrides", ("service_code", "public_url")),
    ):
        write_table(
            workbook.create_sheet(sheet_name),
            table_name=table_name,
            columns=columns,
            table_rows=[],
            banners=banner_rows,
        )

    workbook.save(path)
    _freeze_package_timestamps(path)
    return path


def _seed_event_programme(today: dt.date) -> str:
    """Publish the programme through the real synchronisation path.

    The synthetic workbook is handed to the same `synchronize_public_workbook`
    production uses, so the parser, the contract validation, the import registry,
    the atomic publication and the audit trail are all exercised. No
    `EventProgrammeItem` row is written directly, and nothing here contacts
    OneDrive: the downloader writes the file the caller asked for.
    """
    from apps.event_programme.public_download import XLSX_MIME_TYPE, PublicDownload
    from apps.event_programme.sync import SyncLocked, advisory_lock, synchronize_public_workbook

    def downloader(destination: Path) -> PublicDownload:
        _write_event_programme_workbook(destination, today)
        payload = destination.read_bytes()
        return PublicDownload(
            path=destination,
            size_bytes=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            content_type=XLSX_MIME_TYPE,
            final_host="synthetic-seed.invalid",
        )

    try:
        with advisory_lock():
            outcome = synchronize_public_workbook(downloader=downloader)
    except SyncLocked as error:
        raise CommandError(f"Event-programme seed could not take the lock: {error}") from error
    return f"sündmuste programm: {outcome.result} ({outcome.rows_imported} sündmust)"


# --------------------------------------------------------------------------
# Public Koda.ee feeds — through their own synchronisation services
# --------------------------------------------------------------------------


def _seed_events(today: dt.date) -> str:
    from apps.core.canonical import canonical_checksum
    from apps.events.collector import EventCollection, EventEntry
    from apps.events.sync import synchronize_events

    entries: list[EventEntry] = []

    def add(index: int, *, title: str, starts_on: dt.date, ends_on: dt.date | None = None) -> None:
        entries.append(
            EventEntry(
                stable_key=f"seed-event-{index}",
                title=title,
                canonical_url=f"https://www.koda.ee/et/sundmused/sunteetiline-{index}",
                category=LONG_CATEGORY if index % 3 == 0 else "Sünteetiline koolitus",
                summary="",
                starts_on=starts_on,
                ends_on=ends_on,
                starts_at=None,
                ends_at=None,
                location=LONG_LOCATION if index % 4 == 0 else "Sünteetiline saal",
                source_order=index,
            )
        )

    # The overflow candidate: a very long linked title carrying the visually
    # hidden "(koda.ee, avaneb uuel vahelehel)" suffix.
    add(1, title=LONG_TITLE, starts_on=today + dt.timedelta(days=2))
    # A multi-day range, so the date column renders two dates.
    add(
        2,
        title="Sünteetiline mitmepäevane sündmus",
        starts_on=today + dt.timedelta(days=5),
        ends_on=today + dt.timedelta(days=7),
    )
    # Month and year boundaries, where date formatting most often breaks.
    first_next_month = (today.replace(day=1) + dt.timedelta(days=32)).replace(day=1)
    add(3, title="Sünteetiline kuupiiri sündmus", starts_on=first_next_month - dt.timedelta(days=1))
    add(4, title="Sünteetiline uue kuu sündmus", starts_on=first_next_month)
    add(5, title="Sünteetiline aastavahetuse sündmus", starts_on=dt.date(today.year, 12, 31))
    add(6, title="Sünteetiline uue aasta sündmus", starts_on=dt.date(today.year + 1, 1, 2))
    # Enough further events that the list scrolls.
    for index in range(7, 19):
        add(
            index, title=f"Sünteetiline sündmus {index}", starts_on=today + dt.timedelta(days=index)
        )

    entries.sort(key=lambda item: (item.starts_on, item.title, item.stable_key))
    entries = [
        EventEntry(**{**vars(entry), "source_order": position})
        for position, entry in enumerate(entries)
    ]
    canonical = {
        "dataset": "koda-public-events",
        "schema_version": "1.0",
        "items": [
            {
                "key": entry.stable_key,
                "title": entry.title,
                "url": entry.canonical_url,
                "category": entry.category,
                "starts_on": entry.starts_on,
                "ends_on": entry.ends_on,
                "starts_at": entry.starts_at,
                "ends_at": entry.ends_at,
                "location": entry.location,
            }
            for entry in entries
        ],
    }
    checksum, size = canonical_checksum(canonical)
    collection = EventCollection(
        entries=tuple(entries),
        sha256=checksum,
        size_bytes=size,
        canonical=canonical,
        pages_fetched=1,
        details_fetched=len(entries),
        skipped_non_events=0,
        skipped_past=0,
    )
    outcome = synchronize_events(collector=lambda **_kwargs: collection)
    return f"sündmused: {outcome.result} ({len(entries)} sündmust)"


#: How many synthetic articles the feed publishes. More than one archive page,
#: so `/uudised/` can be tested with a pager on screen.
NEWS_ARTICLES = 40


def _seed_news(today: dt.date) -> str:
    from apps.core.canonical import canonical_checksum
    from apps.news.collector import NewsCollection, NewsEntry
    from apps.news.sync import synchronize_news

    midnight = dt.datetime.combine(today, dt.time(9, 0), tzinfo=dt.UTC)
    entries: list[NewsEntry] = []
    # Deep enough that the news archive has more than one page of thirty, which
    # is the only way a browser test can prove the pager works at all. They are
    # dated a day apart, so every period preset selects a different slice.
    for index in range(1, NEWS_ARTICLES + 1):
        entries.append(
            NewsEntry(
                guid=f"seed-news-{index}",
                title=LONG_TITLE if index == 1 else f"Sünteetiline uudise pealkiri {index}",
                canonical_url=f"https://www.koda.ee/et/uudised/sunteetiline-{index}",
                published_at=midnight - dt.timedelta(days=index),
                category=LONG_CATEGORY if index % 4 == 0 else "Sünteetiline rubriik",
                summary=(
                    "Sünteetiline kokkuvõte, mis on piisavalt pikk, et kontrollida "
                    "teksti murdmist ja kärpimist kaardi laiuses."
                ),
                source_order=index - 1,
            )
        )
    canonical = {
        "dataset": "koda-public-news",
        "schema_version": "1.0",
        "items": [
            {
                "guid": entry.guid,
                "title": entry.title,
                "url": entry.canonical_url,
                "published_at": entry.published_at,
                "category": entry.category,
                "summary": entry.summary,
            }
            for entry in entries
        ],
    }
    checksum, size = canonical_checksum(canonical)
    collection = NewsCollection(
        entries=tuple(entries),
        sha256=checksum,
        size_bytes=size,
        canonical=canonical,
        etag="",
        last_modified="",
    )
    outcome = synchronize_news(collector=lambda **_kwargs: collection)
    return f"uudised: {outcome.result} ({len(entries)} uudist)"


def _seed_public_membership() -> str:
    from apps.core.canonical import canonical_checksum
    from apps.membership.collector import MembershipCollection
    from apps.membership.sync import synchronize_membership

    def collection(total: int) -> MembershipCollection:
        canonical = {
            "dataset": "koda-public-members",
            "schema_version": "1.0",
            "total_members": total,
        }
        checksum, size = canonical_checksum(canonical)
        return MembershipCollection(
            total_members=total,
            sha256=checksum,
            size_bytes=size,
            canonical=canonical,
            etag="",
            last_modified="",
            duplicate_identities=0,
            rejected_rows=0,
        )

    # Two readings, so the directory count has a predecessor to compare with.
    # Both are four-digit, which is what makes grouped-number formatting visible.
    results = []
    for total in (4187, 4203):
        outcome = synchronize_membership(collector=lambda total=total, **_kwargs: collection(total))
        results.append(outcome.result)
    return f"liikmed (avalik): {results[-1]} (4203)"


# --------------------------------------------------------------------------
# Internal board-report history — through the manual publication service
# --------------------------------------------------------------------------


def _seed_internal_membership(today: dt.date) -> str:
    from apps.membership.manual import ManualReport, publish_manual_report
    from apps.membership.quality import MetricFacts

    # Six readings across roughly a year, so both overview trend lines have
    # enough points to draw. Values move plausibly and are all invented.
    plan = [
        (330, 4050, 3810, "1180000.00", "1300000.00", 210, 41, 12),
        (270, 4090, 3860, "1205500.00", "1300000.00", 260, 38, 18),
        (210, 4120, 3905, "1231000.50", "1300000.00", 300, 35, 24),
        (150, 4150, 3950, "1252750.25", "1310000.00", 340, 0, 31),
        (90, 4176, 3988, "1268400.00", "1310000.00", 372, 29, 37),
        (30, 4203, 4025, "1276101.00", "1310000.00", 401, None, 44),
    ]
    # Monthly arrivals, so the recruitment chart has a subject year and an
    # earlier one to draw behind it.
    #
    # A report may only fill months up to its own observation date — a board
    # report cannot state how many joined in a month that has not happened when
    # it was written — so each report carries its own year up to its own month
    # and the series is built across the six of them.
    #
    # February is an explicit `0` and March is left out entirely. The chart has
    # to keep "nobody joined" apart from "nobody reported", and only a seed
    # carrying both shapes can prove that it does.
    latest_date = today - dt.timedelta(days=plan[-1][0])

    def monthly_for(when: dt.date) -> dict[int, int]:
        if when.year < latest_date.year:
            return {number: 18 + number for number in range(1, when.month + 1)}
        return {
            number: (0 if number == 2 else 20 + number)
            for number in range(1, when.month + 1)
            if number != 3
        }

    # Movements and removal reasons ride on the newest report, which is the one
    # the movement section describes. Every band reports both directions except
    # the largest, which reports only arrivals — a band with one side missing
    # must show no net rather than a net that counts a gap as zero.
    #
    # Neither table is marked complete, and that is the honest flag rather than
    # a way around the cross-checks: a table missing one band's departures is a
    # partial table. `publish_manual_report` only reconciles these sums against
    # the year-to-date figures when the report claims completeness, which is the
    # right rule — a partly filled table is an ordinary thing to have and must
    # not be rejected for failing to add up.
    size_joined = {
        "employees_1_4": 21,
        "employees_20_49": 27,
        "employees_100_249": 14,
        "employees_250_499": 8,
    }
    size_removed = {
        "employees_1_4": 38,
        "employees_20_49": 22,
        "employees_100_249": 11,
    }
    reasons = {
        "dissolved_bankrupt_merged_inactive_missing": 120,
        "voluntary_debt_financial_or_other": 84,
        "voluntary_no_service_value": 31,
    }

    published = 0
    for offset, total, paid, received, budget, new_ytd, suspended, removed in plan:
        when = today - dt.timedelta(days=offset)
        is_latest = offset == plan[-1][0]
        report = ManualReport(
            observation_date=when,
            reported_year=when.year,
            document_title="Sünteetiline juhatuse aruanne",
            source_note="Sünteetiline seeme, mitte tegelik aruanne.",
            monthly_year=when.year,
            monthly_new_members=monthly_for(when),
            joined_by_band=size_joined if is_latest else {},
            removed_by_band=size_removed if is_latest else {},
            size_table_complete=False,
            removal_reasons=reasons if is_latest else {},
            reasons_complete=False,
            facts=MetricFacts(
                total_members=total,
                paid_members=paid,
                membership_fees_received_eur=Decimal(received),
                membership_fee_budget_eur=Decimal(budget),
                # Left unreported on purpose: the page then shows the computed
                # percentage and says which basis it used.
                membership_fee_collection_pct_reported=None,
                new_members_ytd=new_ytd,
                # `0` on one reading and `None` on another, so the interface has
                # to distinguish "nobody was suspended" from "nobody counted".
                suspended_members=suspended,
                removed_members_ytd=removed,
            ),
        )
        publish_manual_report(report)
        published += 1
    return f"liikmeskonna aruanded (sisemine): {published} vaatlust"


# --------------------------------------------------------------------------
# Website analytics — through the real GA4 publication path
# --------------------------------------------------------------------------

#: How much history to publish. Long enough that `30 päeva` and `90 päeva` are
#: both offered and the longer windows are visibly disabled, so the browser
#: suite sees an offered control and a refused one rather than only one branch.
ANALYTICS_DAYS = 45

#: Paths that must never reach a content ranking, with the traffic they carry.
#: They are the whole reason the ranking has an exclusion registry: on the real
#: property the language roots alone outweigh every article, so a seed without
#: them cannot show that the registry does anything. Each family in
#: `apps.visibility.content_ranking` is represented once.
ANALYTICS_UTILITY_PAGES = (
    ("/et", 900),
    ("/en", 300),
    ("/ru", 120),
    ("/et/search/node", 260),
    ("/et/cart", 180),
    ("/et/user/login", 60),
    ("/403.html", 90),
    ("/et/node/9001", 70),
)

#: A section's own listing page, which is excluded from the ranking of the
#: content it lists — otherwise every section is topped by its index.
ANALYTICS_INDEX_PAGES = (
    ("/et/uudised", 140),
    ("/et/sundmused", 130),
    ("/et/teenused", 110),
)


#: A path seeded far below the Top 20 on purpose, and the words that find it.
#: Search exists for the page a ranking cannot reach, so proving it works needs
#: a target the ranking never shows — and the term appears in no path, so only
#: the title catalogue can find it.
ANALYTICS_QUIET_PATH = "/et/uudised/sunteetiline-12"

#: How many of the seeded articles carry measured traffic. The rest are
#: catalogued and unmeasured, which is a real and common state.
MEASURED_ARTICLES = 24
ANALYTICS_QUIET_TITLE_TERM = "pealkiri 12"


def _analytics_content_pages() -> tuple[tuple[str, int], ...]:
    """The rankable paths, aligned with what the other seeders publish.

    The paths match `_seed_news` and `_seed_events` exactly, which is what lets
    a row resolve to a real title. Both halves of that are worth having:

    - **news are catalogued**, because `synchronize_news` records every item it
      publishes in `NewsResource`. So news rows show titles, `LONG_TITLE` among
      them — and it is given the heaviest traffic in the section deliberately.
      A very long linked title carrying a visually hidden suffix is the exact
      shape that widened a page by 152 pixels once, and rank one is the only
      place the layout suite will ever measure it;
    - **events and services are not.** `PublicEventResource` is filled by the
      sitemap discovery crawl, not by `synchronize_events`, and services have no
      title catalogue anywhere in the application. Their rows therefore render
      as paths — which is not a gap in the seed but the honest answer for a page
      DashKoda cannot name, and the state a real event page is in until its link
      is backfilled. Having both on screen at once is the point.
    """
    rows: list[tuple[str, int]] = []
    # Deliberately fewer than the seed publishes: the articles beyond this are
    # catalogued but unmeasured, which is what puts a real `—` in the archive's
    # view column and an unmeasured row behind the measured ones when it is
    # ranked. A seed where everything is measured cannot show either.
    for index in range(1, MEASURED_ARTICLES + 1):
        weight = 96 if index == 1 else 90 - index * 3
        # The last article is nearly silent, so it can never drift into the Top
        # 20 and quietly make the search tests assert nothing.
        rows.append((f"/et/uudised/sunteetiline-{index}", 2 if index == 12 else max(weight, 6)))
    for index in range(1, 19):
        weight = 88 if index == 1 else 80 - index * 3
        rows.append((f"/et/sundmused/sunteetiline-{index}", max(weight, 4)))
    for index in range(1, 7):
        rows.append((f"/et/teenused/sunteetiline-teenus-{index}", 70 - index * 4))
    return tuple(rows)


#: Every page row, utility and content together. The site's own figures are the
#: sum of all of them, which is what lets a test show that excluding a path from
#: a *ranking* leaves the website's totals untouched.
ANALYTICS_PAGES = ANALYTICS_UTILITY_PAGES + ANALYTICS_INDEX_PAGES + _analytics_content_pages()

#: Acquisition channels and their share of a day's sessions, in whole percent.
#: The names are GA4's own default channel group, which is not Chamber data —
#: it is the vocabulary the report arrives in.
ANALYTICS_CHANNELS = (
    ("Organic Search", 42),
    ("Direct", 27),
    ("Email", 14),
    ("Organic Social", 11),
    ("Referral", 6),
)


def _analytics_day(report_date: dt.date):
    """One synthetic reporting day, shaped so the chart is not a straight line.

    The weekday rhythm is deterministic, so re-running publishes an identical
    canonical payload and the sync reports `unchanged` rather than filling the
    history with revisions of itself.
    """
    from apps.visibility.ga4 import ChannelRow, DayReading, PageRow

    # Quieter at the weekend. Integer arithmetic throughout: a float would make
    # the canonical payload depend on binary rounding, and the checksum with it.
    scale = 4 if report_date.weekday() >= 5 else 10
    pages = tuple(
        PageRow(path=path, page_views=max(base * scale // 10, 1)) for path, base in ANALYTICS_PAGES
    )

    # The site total *is* the sum of the page rows. Anything else would make
    # "excluded from the list, never from the total" untestable here.
    page_views = sum(row.page_views for row in pages)
    sessions = max(page_views // 3, 1)

    channels: list[ChannelRow] = []
    assigned = 0
    for name, share in ANALYTICS_CHANNELS[1:]:
        count = sessions * share // 100
        assigned += count
        channels.append(ChannelRow(channel=name, sessions=count))
    # The largest channel absorbs the rounding, so the parts always sum to the
    # whole and a reader can add the column up.
    channels.insert(
        0, ChannelRow(channel=ANALYTICS_CHANNELS[0][0], sessions=max(sessions - assigned, 0))
    )

    return DayReading(
        report_date=report_date,
        sessions=sessions,
        active_users=sessions * 8 // 10,
        new_users=sessions * 3 // 10,
        page_views=page_views,
        engaged_sessions=sessions * 5 // 10,
        user_engagement_seconds=sessions * 47,
        pages=pages,
        channels=tuple(channels),
        has_page_detail=True,
        has_channel_detail=True,
    ).validate()


class _SeedGa4Collector:
    """Stands in for the Data API at the seam the real collector uses.

    `synchronize_ga4` takes a collector and owns publication itself, so seeding
    substitutes the transport and nothing else: the same normalisation, the same
    canonical checksum, the same import run, the same immutable revisions. No
    request is made and no property ID or credential is read.
    """

    def collect_range(self, *, start: dt.date, end: dt.date, with_pages=True, with_channels=True):
        from apps.visibility.ga4 import CollectionCounts, RangeCollection

        days = {}
        current = start
        while current <= end:
            days[current] = _analytics_day(current)
            current += dt.timedelta(days=1)
        return RangeCollection(
            days=days,
            counts=CollectionCounts(
                requests=0,
                site_rows=len(days),
                page_rows=sum(len(day.pages) for day in days.values()),
                channel_rows=sum(len(day.channels) for day in days.values()),
            ),
        )


def _seed_website_analytics(today: dt.date) -> str:
    """Publish a synthetic GA4 history so the traffic section exists at all.

    Without it `overview.html` renders the `Lisamisel` empty state and the whole
    website section — the chart, the channel table, the content ranking and the
    page search — is invisible to every browser test. Two defects shipped
    through a fully green suite that way on 2026-08-11: the view dropped the
    `otsing` parameter, and the template hid the search box behind the ranking
    it empties.
    """
    from apps.visibility.ga4_sync import synchronize_ga4

    # `synchronize_ga4` clamps to the last completed day itself; the window is
    # stated in full so the seeded span does not depend on that clamp.
    end = today - dt.timedelta(days=1)
    outcome = synchronize_ga4(
        collector=_SeedGa4Collector(),
        start=end - dt.timedelta(days=ANALYTICS_DAYS - 1),
        end=end,
        today=today,
    )
    return (
        f"veebistatistika: {outcome.result} "
        f"({ANALYTICS_DAYS} päeva, {len(ANALYTICS_PAGES)} lehekülge päevas)"
    )


# --------------------------------------------------------------------------
# Visibility — through the manual submission service
# --------------------------------------------------------------------------


def _seed_visibility(today: dt.date) -> str:
    from apps.visibility.manual import VisibilitySubmission, publish_submission
    from apps.visibility.registry import VisibilityMetric

    # Three readings, so every channel card has a trend and a change to state.
    # The oldest deliberately omits two metrics: a channel nobody has read yet
    # must show "andmed puuduvad" rather than a zero.
    plan = [
        (
            120,
            {
                VisibilityMetric.NEWSLETTER_ETEATAJA: 8120,
                VisibilityMetric.NEWSLETTER_ENEWS: 3040,
                VisibilityMetric.NEWSLETTER_EVESTNIK: 1210,
                VisibilityMetric.FACEBOOK_FOLLOWERS: 5210,
                VisibilityMetric.LINKEDIN_FOLLOWERS: 4130,
            },
        ),
        (
            60,
            {
                VisibilityMetric.NEWSLETTER_ETEATAJA: 8260,
                VisibilityMetric.NEWSLETTER_ENEWS: 3105,
                VisibilityMetric.NEWSLETTER_EVESTNIK: 1188,
                VisibilityMetric.FACEBOOK_FOLLOWERS: 5344,
                VisibilityMetric.LINKEDIN_FOLLOWERS: 4402,
                VisibilityMetric.INSTAGRAM_FOLLOWERS: 1870,
                VisibilityMetric.YOUTUBE_SUBSCRIBERS: 640,
            },
        ),
        (
            7,
            {
                VisibilityMetric.NEWSLETTER_ETEATAJA: 8395,
                VisibilityMetric.NEWSLETTER_ENEWS: 3162,
                VisibilityMetric.NEWSLETTER_EVESTNIK: 1174,
                VisibilityMetric.FACEBOOK_FOLLOWERS: 5498,
                VisibilityMetric.LINKEDIN_FOLLOWERS: 4655,
                VisibilityMetric.INSTAGRAM_FOLLOWERS: 1994,
                VisibilityMetric.YOUTUBE_SUBSCRIBERS: 671,
            },
        ),
    ]
    for offset, values in plan:
        publish_submission(
            VisibilitySubmission(
                observation_date=today - dt.timedelta(days=offset),
                values={str(key): value for key, value in values.items()},
                note="Sünteetiline seeme.",
            )
        )
    return f"nähtavus: {len(plan)} sisestust"


# --------------------------------------------------------------------------


class Command(BaseCommand):
    help = (
        "Publish deterministic synthetic content for the browser acceptance "
        "suite. Refuses to run under production settings."
    )

    def handle(self, *args, **options):
        module = _require_non_production()
        today = timezone.localdate()

        # The legal-work synchronisation owns its own temporary directory and
        # deletes it on every exit path, so no seeded workbook outlives the
        # command.
        lines = [
            _seed_legal_work(today),
            _seed_event_programme(today),
            _seed_events(today),
            _seed_news(today),
            _seed_public_membership(),
            _seed_internal_membership(today),
            _seed_visibility(today),
            # After the news and the events: the page rows resolve their titles
            # from those catalogues, and a ranking seeded first would show paths
            # where the finished page shows titles.
            _seed_website_analytics(today),
        ]

        self.stdout.write(self.style.SUCCESS(f"Sünteetiline seeme ({module}):"))
        for line in lines:
            self.stdout.write(f"  {line}")
